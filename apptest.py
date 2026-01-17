"""
交易數據分析系統 (Trading Analysis System) v2.4
全方位重構版本：
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔧 Bug 修復:
  - 修正 st.number_input 類型錯誤：所有 Filter 統一為浮點數
  - 確保 value, step, min_value, max_value 類型一致

🏆 英雄榜表格優化:
  - 新增 Sharpe Ratio 至所有 Top 20 表格
  - Scalp 門檻改為「Scalp 盈虧金額門檻」
  - 全局 Filter：盈虧、勝率%、Sharpe、MDD% 實時過濾
  - AID 欄位顯式定義為 TextColumn 確保複製功能

👤 Tab 2 個人診斷優化:
  - 搜尋框支援貼上並 .strip() 清理空格
  - 新增提示 caption

📊 視覺佈局重構 (Tab 1):
  - Violin Plot + 風險回報矩陣放大全寬
  - 獲利因子 + 交易風格並排 (st.columns(2))
  - Violin Plot 統計摘要區塊
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO

# ==================== 頁面配置 ====================
st.set_page_config(
    page_title="交易數據分析系統 v2.4", 
    page_icon="📊", 
    layout="wide", 
    initial_sidebar_state="expanded"
)

# ==================== 欄位映射配置 ====================
COLUMN_MAP = {
    'execution_time': 'Execution Time\n交易时间',
    'open_time': 'Open Time\n开仓时间',
    'aid': 'AID\n用户账号',
    'closed_pl': 'Closed P/L\n平仓盈亏',
    'commission': 'Commission\n手续费',
    'swap': 'Swap\n隔夜利息',
    'instrument': 'Instrument\n交易品种',
    'business_type': 'Business Type\n业务类型',
    'action': 'Action\n交易类型',
    'volume': 'Volume\n开仓数量',
    'side': 'Side\n交易方向'
}

STYLE_COLORS = {
    '極短線 (Scalp)': '#E74C3C',
    '短線 (Intraday)': '#F39C12',
    '中線 (Day Trade)': '#3498DB',
    '長線 (Swing)': '#27AE60'
}


# ==================== 數據載入與預處理 ====================
@st.cache_data(show_spinner=False)
def load_and_preprocess(uploaded_files):
    """載入並預處理交易數據"""
    dfs = []
    for uploaded_file in uploaded_files:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            dfs.append(df)
        except Exception as e:
            st.error(f"讀取檔案 {uploaded_file.name} 時發生錯誤: {e}")
            continue
    
    if not dfs:
        return None
    
    df = pd.concat(dfs, ignore_index=True)
    exec_col = COLUMN_MAP['execution_time']
    if exec_col in df.columns:
        df = df[df[exec_col] != 'Total'].copy()
    df = df.drop_duplicates()
    
    for col in ['execution_time', 'open_time']:
        if COLUMN_MAP[col] in df.columns:
            df[COLUMN_MAP[col]] = pd.to_datetime(df[COLUMN_MAP[col]], errors='coerce')
    
    for col in ['closed_pl', 'commission', 'swap']:
        if COLUMN_MAP[col] in df.columns:
            df[COLUMN_MAP[col]] = df[COLUMN_MAP[col]].fillna(0)
    
    df['Net_PL'] = df[COLUMN_MAP['closed_pl']] + df[COLUMN_MAP['commission']] + df[COLUMN_MAP['swap']]
    
    exec_time = df[COLUMN_MAP['execution_time']]
    open_time = df[COLUMN_MAP['open_time']]
    df['Hold_Seconds'] = np.where(
        pd.notna(exec_time) & pd.notna(open_time), 
        (exec_time - open_time).dt.total_seconds(), 
        np.nan
    )
    df['Hold_Minutes'] = df['Hold_Seconds'] / 60
    
    if COLUMN_MAP['aid'] in df.columns:
        df[COLUMN_MAP['aid']] = (
            df[COLUMN_MAP['aid']]
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.replace(',', '', regex=False)
            .str.strip()
        )
    
    return df


def filter_closing_trades(df):
    """過濾平倉交易"""
    action_col = COLUMN_MAP['action']
    if action_col in df.columns:
        return df[df[action_col] == 'CLOSING'].copy()
    return df


def classify_trading_style(hold_minutes):
    """分類交易風格"""
    if pd.isna(hold_minutes):
        return '短線 (Intraday)'
    elif hold_minutes < 5:
        return '極短線 (Scalp)'
    elif hold_minutes < 60:
        return '短線 (Intraday)'
    elif hold_minutes < 1440:
        return '中線 (Day Trade)'
    else:
        return '長線 (Swing)'


# ==================== 統一英雄榜計算函數 ====================
def calculate_hero_metrics(data_df, initial_balance, scalper_threshold_seconds, 
                           filter_positive=True, min_scalp_pct=None, min_scalp_pl=None,
                           min_pnl=None, min_winrate=None, min_sharpe=None, max_mdd=None):
    """
    統一計算英雄榜指標
    包含：AID | 盈虧 | Scalp盈虧 | Scalp% | Sharpe | Q1 | Median | Q3 | IQR | P. Exp | PF | Rec.F | MDD% | 勝率% | 筆數
    """
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    closing_df = filter_closing_trades(data_df)
    results = []
    
    for aid in closing_df[aid_col].unique():
        aid_data = closing_df[closing_df[aid_col] == aid].copy()
        
        net_pl = aid_data['Net_PL'].sum()
        trade_count = len(aid_data)
        
        if trade_count == 0:
            continue
        
        # 篩選條件：僅正盈虧
        if filter_positive and net_pl <= 0:
            continue
        
        # Scalp 數據
        scalp_trades = aid_data[aid_data['Hold_Seconds'] < scalper_threshold_seconds]
        scalp_count = len(scalp_trades)
        scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
        scalp_pct = (scalp_count / trade_count * 100) if trade_count > 0 else 0
        
        # 勝率
        wins = (aid_data['Net_PL'] > 0).sum()
        losses = trade_count - wins
        win_rate = (wins / trade_count * 100) if trade_count > 0 else 0
        
        # Sharpe Ratio
        if trade_count >= 3:
            mean_pl = aid_data['Net_PL'].mean()
            std_pl = aid_data['Net_PL'].std()
            sharpe = mean_pl / std_pl if std_pl > 0 else 0.0
        else:
            sharpe = 0.0
        
        # MDD% 計算
        aid_sorted = aid_data.sort_values(exec_col)
        if len(aid_sorted) >= 2:
            cumulative_pl = aid_sorted['Net_PL'].cumsum()
            equity = initial_balance + cumulative_pl
            running_max = equity.cummax()
            drawdown = np.where(running_max != 0, (equity - running_max) / running_max * 100, 0)
            mdd_pct = abs(np.min(drawdown))
            max_dd_abs = abs((equity - running_max).min())
        else:
            mdd_pct = 0.0
            max_dd_abs = 0.0
        
        # 應用過濾器（確保類型比較正確）
        if min_scalp_pct is not None and scalp_pct < float(min_scalp_pct):
            continue
        if min_scalp_pl is not None and scalp_pl < float(min_scalp_pl):
            continue
        if min_pnl is not None and net_pl < float(min_pnl):
            continue
        if min_winrate is not None and win_rate < float(min_winrate):
            continue
        if min_sharpe is not None and sharpe < float(min_sharpe):
            continue
        if max_mdd is not None and mdd_pct > float(max_mdd):
            continue
        
        # Box Plot 指標
        q1 = aid_data['Net_PL'].quantile(0.25)
        median = aid_data['Net_PL'].median()
        q3 = aid_data['Net_PL'].quantile(0.75)
        iqr = q3 - q1
        
        # Profit Expectancy
        win_trades = aid_data[aid_data['Net_PL'] > 0]['Net_PL']
        loss_trades = aid_data[aid_data['Net_PL'] < 0]['Net_PL']
        avg_win = win_trades.mean() if len(win_trades) > 0 else 0
        avg_loss = abs(loss_trades.mean()) if len(loss_trades) > 0 else 0
        win_prob = wins / trade_count if trade_count > 0 else 0
        loss_prob = losses / trade_count if trade_count > 0 else 0
        p_exp = (win_prob * avg_win) - (loss_prob * avg_loss)
        
        # Profit Factor
        gains = win_trades.sum() if len(win_trades) > 0 else 0
        total_losses = abs(loss_trades.sum()) if len(loss_trades) > 0 else 0
        pf = gains / total_losses if total_losses > 0 else (5.0 if gains > 0 else 0.0)
        
        # Recovery Factor
        rec_f = net_pl / max_dd_abs if max_dd_abs > 0 else (net_pl if net_pl > 0 else 0.0)
        
        results.append({
            'AID': str(aid),
            '盈虧': round(net_pl, 2),
            'Scalp盈虧': round(scalp_pl, 2),
            'Scalp%': round(scalp_pct, 2),
            'Sharpe': round(sharpe, 2),
            'MDD%': round(mdd_pct, 2),
            'Q1': round(q1, 2),
            'Median': round(median, 2),
            'Q3': round(q3, 2),
            'IQR': round(iqr, 2),
            'P. Exp': round(p_exp, 2),
            'PF': round(pf, 2),
            'Rec.F': round(rec_f, 2),
            '勝率%': round(win_rate, 2),
            '筆數': trade_count
        })
    
    result_df = pd.DataFrame(results)
    if not result_df.empty:
        result_df = result_df.sort_values('盈虧', ascending=False).head(20)
    return result_df


def format_hero_table_display(hero_df):
    """格式化英雄榜表格顯示"""
    if hero_df.empty:
        return hero_df
    
    display_df = hero_df.copy()
    
    # Scalp% emoji
    display_df['Scalp%'] = display_df['Scalp%'].apply(
        lambda x: f"🔥{x:.1f}%" if x > 80 else f"{x:.1f}%"
    )
    
    # Sharpe 顏色
    display_df['Sharpe'] = display_df['Sharpe'].apply(
        lambda x: f"⭐{x:.2f}" if x > 2 else f"{x:.2f}"
    )
    
    # MDD% 紅色警示
    display_df['MDD%'] = display_df['MDD%'].apply(
        lambda x: f"🔴{x:.1f}%" if x > 20 else f"{x:.1f}%"
    )
    
    # P.Exp 顏色
    display_df['P. Exp'] = display_df['P. Exp'].apply(
        lambda x: f"🟢{x:.2f}" if x > 0 else f"🔴{x:.2f}"
    )
    
    # 金額格式
    for col in ['盈虧', 'Scalp盈虧', 'Q1', 'Median', 'Q3', 'IQR']:
        display_df[col] = display_df[col].apply(lambda x: f"${x:,.2f}")
    
    return display_df


def get_table_column_config():
    """
    獲取統一的表格欄位配置 - 確保 AID 為純文字可複製
    🔧 修復：顯式定義 AID 為 TextColumn 確保出現複製小圖示
    """
    return {
        'AID': st.column_config.TextColumn(
            'AID', 
            help='📋 點擊單元格可選取複製', 
            width='small'
        ),
        '盈虧': st.column_config.TextColumn('盈虧', width='medium'),
        'Scalp盈虧': st.column_config.TextColumn('Scalp盈虧', width='medium'),
        'Scalp%': st.column_config.TextColumn('Scalp%', width='small'),
        'Sharpe': st.column_config.TextColumn('Sharpe', width='small'),
        'MDD%': st.column_config.TextColumn('MDD%', width='small'),
        'Q1': st.column_config.TextColumn('Q1', width='small'),
        'Median': st.column_config.TextColumn('Median', width='small'),
        'Q3': st.column_config.TextColumn('Q3', width='small'),
        'IQR': st.column_config.TextColumn('IQR', width='small'),
        'P. Exp': st.column_config.TextColumn('P.Exp', width='small'),
        'PF': st.column_config.NumberColumn('PF', format='%.2f', width='small'),
        'Rec.F': st.column_config.NumberColumn('Rec.F', format='%.2f', width='small'),
        '勝率%': st.column_config.NumberColumn('勝率%', format='%.1f%%', width='small'),
        '筆數': st.column_config.NumberColumn('筆數', format='%d', width='small')
    }


def render_global_filters(key_prefix, default_pnl=0.0, default_winrate=0.0, 
                          default_sharpe=-10.0, default_mdd=100.0):
    """
    渲染全局過濾器
    🔧 修復：確保所有 value, step, min_value, max_value 為統一浮點數類型
    """
    st.markdown("#### 🔧 全局過濾器")
    f1, f2, f3, f4 = st.columns(4)
    
    with f1:
        min_pnl = st.number_input(
            "最低盈虧 ($)", 
            value=float(default_pnl), 
            step=100.0,  # 確保為 float
            key=f"{key_prefix}_pnl", 
            help="僅顯示盈虧 ≥ 此值的客戶"
        )
    with f2:
        min_winrate = st.number_input(
            "最低勝率 (%)", 
            value=float(default_winrate), 
            min_value=0.0,  # 確保為 float
            max_value=100.0,  # 確保為 float
            step=5.0,  # 確保為 float
            key=f"{key_prefix}_wr", 
            help="僅顯示勝率 ≥ 此值的客戶"
        )
    with f3:
        min_sharpe = st.number_input(
            "最低 Sharpe", 
            value=float(default_sharpe), 
            step=0.5,  # 確保為 float
            key=f"{key_prefix}_sharpe", 
            help="僅顯示 Sharpe ≥ 此值的客戶"
        )
    with f4:
        max_mdd = st.number_input(
            "最高 MDD (%)", 
            value=float(default_mdd), 
            min_value=0.0,  # 確保為 float
            max_value=100.0,  # 確保為 float
            step=5.0,  # 確保為 float
            key=f"{key_prefix}_mdd", 
            help="僅顯示 MDD ≤ 此值的客戶"
        )
    
    return min_pnl, min_winrate, min_sharpe, max_mdd


def render_scalper_filters(key_prefix, default_scalp_pct=80.0, default_scalp_pl=0.0):
    """
    渲染 Scalper 專用過濾器
    🔧 修復：「Scalp 盈虧金額門檻」取代舊版本
    """
    s1, s2 = st.columns(2)
    with s1:
        min_scalp_pct = st.slider(
            "Scalp% 門檻", 
            min_value=50, 
            max_value=100, 
            value=int(default_scalp_pct),  # slider 需要 int
            step=5, 
            key=f"{key_prefix}_spct", 
            help="Scalp 交易筆數佔比"
        )
    with s2:
        min_scalp_pl = st.number_input(
            "Scalp 盈虧金額門檻 ($)", 
            value=float(default_scalp_pl),  # 確保為 float
            step=100.0,  # 確保為 float
            key=f"{key_prefix}_spl", 
            help="僅顯示 Scalp 盈虧 ≥ 此值的客戶"
        )
    return float(min_scalp_pct), min_scalp_pl


# ==================== 產品堆疊柱狀圖 ====================
def calculate_product_scalp_breakdown(day_df, scalper_threshold_seconds):
    """計算產品 Scalp 拆解"""
    instrument_col = COLUMN_MAP['instrument']
    closing_df = filter_closing_trades(day_df)
    
    if instrument_col not in closing_df.columns:
        return None, None
    
    results = []
    for product in closing_df[instrument_col].unique():
        prod_data = closing_df[closing_df[instrument_col] == product]
        total_pl = prod_data['Net_PL'].sum()
        scalp_trades = prod_data[prod_data['Hold_Seconds'] < scalper_threshold_seconds]
        non_scalp_trades = prod_data[prod_data['Hold_Seconds'] >= scalper_threshold_seconds]
        scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
        non_scalp_pl = non_scalp_trades['Net_PL'].sum() if not non_scalp_trades.empty else 0
        scalp_pct = (len(scalp_trades) / len(prod_data) * 100) if len(prod_data) > 0 else 0
        
        results.append({
            'Product': product, 
            'Total_PL': total_pl, 
            'Scalp_PL': scalp_pl, 
            'NonScalp_PL': non_scalp_pl, 
            'Scalp_Pct': scalp_pct
        })
    
    result_df = pd.DataFrame(results)
    profit_products = result_df[result_df['Total_PL'] > 0].nlargest(5, 'Total_PL')
    loss_products = result_df[result_df['Total_PL'] < 0].nsmallest(5, 'Total_PL')
    
    return profit_products, loss_products


def create_stacked_product_chart(product_df, is_profit=True):
    """創建堆疊產品柱狀圖"""
    if product_df is None or product_df.empty:
        return None
    
    df = product_df.copy()
    if is_profit:
        non_scalp_color, scalp_color = '#1E8449', '#82E0AA'
        title = '📈 當日盈利產品 Top 5'
    else:
        non_scalp_color, scalp_color = '#922B21', '#F1948A'
        title = '📉 當日虧損產品 Top 5'
    
    df = df.sort_values('Total_PL', ascending=not is_profit)
    
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=df['Product'], 
        x=df['NonScalp_PL'], 
        name='Non-Scalp', 
        orientation='h', 
        marker_color=non_scalp_color, 
        text=df['NonScalp_PL'].apply(lambda x: f"${x:,.0f}"), 
        textposition='inside'
    ))
    fig.add_trace(go.Bar(
        y=df['Product'], 
        x=df['Scalp_PL'], 
        name='Scalp', 
        orientation='h', 
        marker_color=scalp_color, 
        text=df['Scalp_PL'].apply(lambda x: f"${x:,.0f}"), 
        textposition='inside'
    ))
    fig.update_layout(
        title=title, 
        barmode='relative', 
        xaxis_title='盈虧金額 ($)', 
        height=300, 
        legend=dict(orientation="h", y=1.1), 
        plot_bgcolor='rgba(248,249,250,1)'
    )
    fig.add_vline(x=0, line_color="black", line_width=1)
    
    return fig


# ==================== 基本統計計算 ====================
def calculate_all_aid_stats_realtime(df, initial_balance, scalper_threshold_seconds):
    """計算所有 AID 的即時統計"""
    aid_col = COLUMN_MAP['aid']
    volume_col = COLUMN_MAP['volume']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    closing_df = filter_closing_trades(df)
    results = []
    
    for aid in closing_df[aid_col].unique():
        aid_data = closing_df[closing_df[aid_col] == aid].copy()
        
        net_pl = aid_data['Net_PL'].sum()
        trade_count = len(aid_data)
        trade_volume = aid_data[volume_col].sum() if volume_col in aid_data.columns else trade_count
        
        wins = (aid_data['Net_PL'] > 0).sum()
        win_rate = (wins / trade_count * 100) if trade_count > 0 else 0
        
        avg_hold_seconds = aid_data['Hold_Seconds'].mean() if 'Hold_Seconds' in aid_data.columns else 0
        avg_hold_seconds = avg_hold_seconds if pd.notna(avg_hold_seconds) else 0
        
        scalper_trades = aid_data[aid_data['Hold_Seconds'] < scalper_threshold_seconds]
        scalper_count = len(scalper_trades)
        scalper_ratio = (scalper_count / trade_count * 100) if trade_count > 0 else 0
        scalper_pl = scalper_trades['Net_PL'].sum() if not scalper_trades.empty else 0
        
        q1 = aid_data['Net_PL'].quantile(0.25)
        median = aid_data['Net_PL'].median()
        q3 = aid_data['Net_PL'].quantile(0.75)
        
        # Sharpe
        if trade_count >= 3:
            sharpe = aid_data['Net_PL'].mean() / aid_data['Net_PL'].std() if aid_data['Net_PL'].std() > 0 else 0
        else:
            sharpe = 0
        
        aid_sorted = aid_data.sort_values(exec_col)
        if len(aid_sorted) >= 2:
            cumulative_pl = aid_sorted['Net_PL'].cumsum()
            equity = initial_balance + cumulative_pl
            running_max = equity.cummax()
            drawdown = np.where(running_max != 0, (equity - running_max) / running_max, 0)
            mdd_pct = abs(np.min(drawdown) * 100)
        else:
            mdd_pct = 0.0
        
        gains = aid_data[aid_data[closed_pl_col] > 0][closed_pl_col].sum()
        losses = abs(aid_data[aid_data[closed_pl_col] < 0][closed_pl_col].sum())
        profit_factor = gains / losses if losses > 0 else (5.0 if gains > 0 else 0)
        
        if instrument_col in aid_data.columns and not aid_data[instrument_col].empty:
            main_symbol = aid_data[instrument_col].mode().iloc[0] if len(aid_data[instrument_col].mode()) > 0 else 'N/A'
        else:
            main_symbol = 'N/A'
        
        results.append({
            'AID': str(aid), 
            'Net_PL': round(net_pl, 2), 
            'Trade_Count': trade_count,
            'Trade_Volume': round(trade_volume, 2), 
            'Win_Rate': round(win_rate, 2),
            'Avg_Hold_Seconds': round(avg_hold_seconds, 2), 
            'MDD_Pct': round(mdd_pct, 2),
            'Profit_Factor': round(profit_factor, 2), 
            'Scalper_Count': scalper_count,
            'Scalper_Ratio': round(scalper_ratio, 2), 
            'Scalper_PL': round(scalper_pl, 2),
            'Sharpe': round(sharpe, 2),
            'Q1': round(q1, 2), 
            'Median': round(median, 2), 
            'Q3': round(q3, 2),
            'Main_Symbol': main_symbol
        })
    
    return pd.DataFrame(results)


# ==================== 深度行為分析 ====================
def calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds):
    """計算深度行為統計"""
    side_col = COLUMN_MAP['side']
    
    total_trades = len(client_df)
    total_pl = client_df['Net_PL'].sum()
    total_minutes = client_df['Hold_Minutes'].sum() if 'Hold_Minutes' in client_df.columns else 0
    total_minutes = total_minutes if pd.notna(total_minutes) else 0
    
    pnl_signs = (client_df['Net_PL'] > 0).astype(int)
    streaks = []
    current_streak = 1
    current_type = pnl_signs.iloc[0] if len(pnl_signs) > 0 else 0
    
    for i in range(1, len(pnl_signs)):
        if pnl_signs.iloc[i] == current_type:
            current_streak += 1
        else:
            streaks.append((current_type, current_streak))
            current_streak = 1
            current_type = pnl_signs.iloc[i]
    if len(pnl_signs) > 0:
        streaks.append((current_type, current_streak))
    
    win_streaks = [s[1] for s in streaks if s[0] == 1]
    loss_streaks = [s[1] for s in streaks if s[0] == 0]
    max_win_streak = max(win_streaks) if win_streaks else 0
    max_loss_streak = max(loss_streaks) if loss_streaks else 0
    
    client_sorted = client_df.sort_values(COLUMN_MAP['execution_time']).copy()
    client_sorted['streak_group'] = (client_sorted['Net_PL'] > 0).ne((client_sorted['Net_PL'] > 0).shift()).cumsum()
    streak_sums = client_sorted.groupby('streak_group')['Net_PL'].sum()
    max_streak_profit = streak_sums.max() if not streak_sums.empty else 0
    max_streak_loss = streak_sums.min() if not streak_sums.empty else 0
    
    buy_trades = client_df[client_df[side_col] == 'BUY'] if side_col in client_df.columns else pd.DataFrame()
    sell_trades = client_df[client_df[side_col] == 'SELL'] if side_col in client_df.columns else pd.DataFrame()
    
    buy_count, sell_count = len(buy_trades), len(sell_trades)
    buy_ratio = (buy_count / total_trades * 100) if total_trades > 0 else 0
    sell_ratio = (sell_count / total_trades * 100) if total_trades > 0 else 0
    buy_pl = buy_trades['Net_PL'].sum() if not buy_trades.empty else 0
    sell_pl = sell_trades['Net_PL'].sum() if not sell_trades.empty else 0
    buy_wins = (buy_trades['Net_PL'] > 0).sum() if not buy_trades.empty else 0
    sell_wins = (sell_trades['Net_PL'] > 0).sum() if not sell_trades.empty else 0
    buy_winrate = (buy_wins / buy_count * 100) if buy_count > 0 else 0
    sell_winrate = (sell_wins / sell_count * 100) if sell_count > 0 else 0
    
    scalp_trades = client_df[client_df['Hold_Seconds'] < scalper_threshold_seconds]
    scalp_count = len(scalp_trades)
    scalp_ratio = (scalp_count / total_trades * 100) if total_trades > 0 else 0
    scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
    scalp_contribution = (scalp_pl / total_pl * 100) if total_pl != 0 else 0
    scalp_wins = (scalp_trades['Net_PL'] > 0).sum() if not scalp_trades.empty else 0
    scalp_winrate = (scalp_wins / scalp_count * 100) if scalp_count > 0 else 0
    
    q1 = client_df['Net_PL'].quantile(0.25)
    median = client_df['Net_PL'].median()
    q3 = client_df['Net_PL'].quantile(0.75)
    iqr = q3 - q1
    
    avg_minutes = total_minutes / total_trades if total_trades > 0 else 0
    profit_per_minute = total_pl / total_minutes if total_minutes > 0 else 0
    avg_seconds = avg_minutes * 60
    hours = int(avg_seconds // 3600)
    minutes = int((avg_seconds % 3600) // 60)
    seconds = int(avg_seconds % 60)
    avg_hold_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    avg_hold_days = avg_minutes / 1440
    
    return {
        'max_win_streak': max_win_streak, 
        'max_loss_streak': max_loss_streak,
        'max_streak_profit': max_streak_profit, 
        'max_streak_loss': max_streak_loss,
        'buy_count': buy_count, 
        'sell_count': sell_count,
        'buy_ratio': buy_ratio, 
        'sell_ratio': sell_ratio,
        'buy_pl': buy_pl, 
        'sell_pl': sell_pl,
        'buy_winrate': buy_winrate, 
        'sell_winrate': sell_winrate,
        'scalp_count': scalp_count, 
        'scalp_ratio': scalp_ratio,
        'scalp_pl': scalp_pl, 
        'scalp_contribution': scalp_contribution,
        'scalp_winrate': scalp_winrate, 
        'avg_hold_formatted': avg_hold_formatted,
        'avg_hold_days': avg_hold_days, 
        'profit_per_minute': profit_per_minute,
        'q1': q1, 
        'median': median, 
        'q3': q3, 
        'iqr': iqr
    }


# ==================== 圖表函數 ====================
def create_cumulative_pnl_chart(df, initial_balance, scalper_threshold_seconds):
    """創建累計盈虧走勢圖"""
    exec_col = COLUMN_MAP['execution_time']
    scalper_minutes = scalper_threshold_seconds / 60
    
    closing_df = filter_closing_trades(df)
    df_sorted = closing_df.sort_values(exec_col).copy()
    df_sorted['Date'] = df_sorted[exec_col].dt.date
    
    daily_pnl = df_sorted.groupby('Date')['Net_PL'].sum().reset_index()
    daily_pnl.columns = ['Date', 'Daily_PL']
    daily_pnl = daily_pnl.sort_values('Date')
    daily_pnl['Cumulative_PL'] = daily_pnl['Daily_PL'].cumsum()
    
    scalper_df = df_sorted[df_sorted['Hold_Seconds'] < scalper_threshold_seconds].copy()
    if not scalper_df.empty:
        scalper_daily_pnl = scalper_df.groupby('Date')['Net_PL'].sum().reset_index()
        scalper_daily_pnl.columns = ['Date', 'Scalper_Daily_PL']
    else:
        scalper_daily_pnl = pd.DataFrame({'Date': daily_pnl['Date'], 'Scalper_Daily_PL': 0})
    
    merged_df = daily_pnl.merge(scalper_daily_pnl, on='Date', how='left')
    merged_df['Scalper_Daily_PL'] = merged_df['Scalper_Daily_PL'].fillna(0)
    merged_df['Scalper_Cumulative_PL'] = merged_df['Scalper_Daily_PL'].cumsum()
    merged_df['Date'] = pd.to_datetime(merged_df['Date'])
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=merged_df['Date'], 
        y=merged_df['Cumulative_PL'], 
        mode='lines+markers', 
        name='整體累計', 
        line=dict(color='#2E86AB', width=2.5)
    ))
    fig.add_trace(go.Scatter(
        x=merged_df['Date'], 
        y=merged_df['Scalper_Cumulative_PL'], 
        mode='lines+markers', 
        name=f'Scalper (<{scalper_minutes:.0f}分鐘)', 
        line=dict(color='#F39C12', width=2.5, dash='dot')
    ))
    fig.add_hline(y=0, line_dash="dash", line_color="gray", line_width=1.5)
    fig.update_layout(
        title='📈 累計淨盈虧走勢', 
        xaxis_title='日期', 
        yaxis_title='累計淨盈虧 ($)', 
        height=450, 
        legend=dict(orientation="h", y=1.02), 
        plot_bgcolor='rgba(248,249,250,1)'
    )
    
    return fig, {
        'total_pnl': merged_df['Cumulative_PL'].iloc[-1] if len(merged_df) > 0 else 0, 
        'scalper_pnl': merged_df['Scalper_Cumulative_PL'].iloc[-1] if len(merged_df) > 0 else 0
    }


def create_violin_plot_with_stats(df):
    """
    創建小提琴圖並返回統計數據
    📊 放大顯示 + 統計摘要
    """
    aid_col = COLUMN_MAP['aid']
    closing_df = filter_closing_trades(df)
    aid_pl = closing_df.groupby(aid_col)['Net_PL'].sum().reset_index()
    aid_pl.columns = ['AID', 'Net_PL']
    
    # 統計數據
    stats = {
        'count': len(aid_pl),
        'mean': aid_pl['Net_PL'].mean(),
        'median': aid_pl['Net_PL'].median(),
        'std': aid_pl['Net_PL'].std(),
        'q1': aid_pl['Net_PL'].quantile(0.25),
        'q3': aid_pl['Net_PL'].quantile(0.75),
        'min': aid_pl['Net_PL'].min(),
        'max': aid_pl['Net_PL'].max(),
        'profitable': (aid_pl['Net_PL'] > 0).sum(),
        'losing': (aid_pl['Net_PL'] <= 0).sum()
    }
    stats['iqr'] = stats['q3'] - stats['q1']
    stats['lower_fence'] = stats['q1'] - 1.5 * stats['iqr']
    stats['upper_fence'] = stats['q3'] + 1.5 * stats['iqr']
    stats['outliers'] = len(aid_pl[
        (aid_pl['Net_PL'] < stats['lower_fence']) | 
        (aid_pl['Net_PL'] > stats['upper_fence'])
    ])
    
    Q1_pct = aid_pl['Net_PL'].quantile(0.01)
    Q99_pct = aid_pl['Net_PL'].quantile(0.99)
    
    fig = go.Figure()
    fig.add_trace(go.Violin(
        x=aid_pl['Net_PL'], 
        y=['盈虧分布'] * len(aid_pl), 
        orientation='h',
        box_visible=True, 
        meanline_visible=True, 
        line_color='#2C3E50',
        fillcolor='rgba(52, 152, 219, 0.5)', 
        points='all', 
        pointpos=-0.5, 
        jitter=0.3,
        marker=dict(color='#3498DB', size=6, opacity=0.6),
        customdata=aid_pl['AID'].values,
        hovertemplate='<b>AID:</b> %{customdata}<br><b>Net_PL:</b> $%{x:,.2f}<extra></extra>'
    ))
    
    x_padding = (Q99_pct - Q1_pct) * 0.1
    fig.add_vline(x=0, line_color="black", line_width=3)
    fig.update_layout(
        title='🎻 客戶盈虧分佈 (Violin Plot)',
        height=750,  # 放大高度
        xaxis=dict(title='累計淨盈虧 ($)', range=[Q1_pct - x_padding, Q99_pct + x_padding]),
        yaxis=dict(showticklabels=False),
        plot_bgcolor='rgba(248,249,250,1)'
    )
    
    return fig, stats


def create_trading_style_pie(df, title="交易風格分佈"):
    """創建交易風格圓餅圖"""
    closing_df = filter_closing_trades(df)
    if 'Hold_Minutes' not in closing_df.columns or closing_df['Hold_Minutes'].isna().all():
        return None
    
    closing_df = closing_df.copy()
    closing_df['Trading_Style'] = closing_df['Hold_Minutes'].apply(classify_trading_style)
    style_counts = closing_df['Trading_Style'].value_counts().reset_index()
    style_counts.columns = ['風格', '筆數']
    
    fig = px.pie(
        style_counts, 
        values='筆數', 
        names='風格', 
        hole=0.4, 
        color='風格', 
        color_discrete_map=STYLE_COLORS, 
        title=title
    )
    fig.update_traces(textposition='inside', textinfo='label+percent')
    fig.update_layout(height=400, legend=dict(orientation="h", y=-0.15))
    return fig


def create_profit_factor_chart_colored(aid_stats_df):
    """創建獲利因子分佈圖"""
    pf_data = aid_stats_df[['AID', 'Profit_Factor', 'Net_PL', 'Trade_Count']].copy()
    pf_display = pf_data[pf_data['Profit_Factor'] <= 5].copy()
    
    bins = [0, 0.5, 1.0, 1.5, 2.0, 3.0, 5.0]
    pf_display['PF_Bin'] = pd.cut(pf_display['Profit_Factor'], bins=bins, right=False)
    bin_stats = pf_display.groupby('PF_Bin', observed=True).size().reset_index(name='Count')
    bin_stats['PF_Bin_Str'] = bin_stats['PF_Bin'].astype(str)
    bin_stats['Color'] = bin_stats['PF_Bin'].apply(
        lambda x: '#E74C3C' if x.right <= 1.0 else '#27AE60'
    )
    
    fig = go.Figure()
    for _, row in bin_stats.iterrows():
        fig.add_trace(go.Bar(
            x=[row['PF_Bin_Str']], 
            y=[row['Count']], 
            marker=dict(color=row['Color'], opacity=0.75), 
            showlegend=False
        ))
    
    fig.add_vline(x=1.5, line_dash="dash", line_color="red", line_width=2, annotation_text="PF=1.0")
    fig.update_layout(
        title='📊 獲利因子分布', 
        xaxis=dict(title='Profit Factor', tickangle=-45), 
        yaxis_title='交易者數', 
        height=400, 
        plot_bgcolor='rgba(248,249,250,1)'
    )
    
    profitable_ratio = (pf_data['Profit_Factor'] > 1.0).sum() / len(pf_data) * 100 if len(pf_data) > 0 else 0
    return fig, profitable_ratio


def create_risk_return_scatter(aid_stats_df, initial_balance):
    """
    創建風險回報矩陣散佈圖
    📊 放大全寬顯示
    """
    scatter_df = aid_stats_df.copy()
    min_size, max_size = 10, 50
    if scatter_df['Trade_Volume'].max() > scatter_df['Trade_Volume'].min():
        scatter_df['Size'] = min_size + (
            (scatter_df['Trade_Volume'] - scatter_df['Trade_Volume'].min()) / 
            (scatter_df['Trade_Volume'].max() - scatter_df['Trade_Volume'].min()) * 
            (max_size - min_size)
        )
    else:
        scatter_df['Size'] = 20
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=scatter_df['MDD_Pct'], 
        y=scatter_df['Net_PL'], 
        mode='markers',
        marker=dict(
            size=scatter_df['Size'], 
            color=scatter_df['Net_PL'], 
            colorscale=['#E74C3C', '#F39C12', '#27AE60'], 
            showscale=True, 
            colorbar=dict(title='盈虧')
        ),
        customdata=np.column_stack((
            scatter_df['AID'], 
            scatter_df['Win_Rate'], 
            scatter_df['Sharpe']
        )),
        hovertemplate=(
            '<b>AID:</b> %{customdata[0]}<br>'
            '<b>淨盈虧:</b> $%{y:,.2f}<br>'
            '<b>MDD:</b> %{x:.1f}%<br>'
            '<b>勝率:</b> %{customdata[1]:.1f}%<br>'
            '<b>Sharpe:</b> %{customdata[2]:.2f}<extra></extra>'
        )
    ))
    fig.update_layout(
        title=f'🎯 風險回報矩陣 (初始資金: ${initial_balance:,})', 
        xaxis=dict(title='MDD (%)', range=[0, 100]), 
        yaxis_title='總盈虧 ($)', 
        height=750,  # 放大高度
        plot_bgcolor='rgba(248,249,250,1)'
    )
    fig.add_hline(y=0, line_dash="dash", line_color="gray")
    fig.add_vline(x=50, line_dash="dash", line_color="gray")
    
    # 象限標註
    fig.add_annotation(x=10, y=0.95, xref="x", yref="paper", text="🌟 低風險高回報", showarrow=False, font=dict(size=12, color="green"))
    fig.add_annotation(x=90, y=0.95, xref="x", yref="paper", text="⚡ 高風險高回報", showarrow=False, font=dict(size=12, color="orange"))
    fig.add_annotation(x=10, y=0.05, xref="x", yref="paper", text="🐢 低風險低回報", showarrow=False, font=dict(size=12, color="gray"))
    fig.add_annotation(x=90, y=0.05, xref="x", yref="paper", text="⚠️ 高風險虧損", showarrow=False, font=dict(size=12, color="red"))
    
    return fig


def create_daily_pnl_chart(df):
    """創建每日盈虧柱狀圖"""
    exec_col = COLUMN_MAP['execution_time']
    closing_df = filter_closing_trades(df)
    df_daily = closing_df.copy()
    df_daily['Date'] = df_daily[exec_col].dt.date
    daily_pnl = df_daily.groupby('Date')['Net_PL'].sum().reset_index()
    daily_pnl.columns = ['日期', '每日盈虧']
    colors = ['#27AE60' if x > 0 else '#E74C3C' for x in daily_pnl['每日盈虧']]
    
    fig = go.Figure()
    fig.add_trace(go.Bar(x=daily_pnl['日期'], y=daily_pnl['每日盈虧'], marker_color=colors))
    fig.add_hline(y=0, line_color="black", line_width=1)
    fig.update_layout(
        title='📅 每日盈虧', 
        xaxis_title='日期', 
        yaxis_title='淨盈虧 ($)', 
        height=350, 
        plot_bgcolor='rgba(248,249,250,1)'
    )
    return fig


def create_client_cumulative_chart(cumulative_df, scalper_minutes):
    """創建個人累計盈虧圖"""
    exec_col = COLUMN_MAP['execution_time']
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=cumulative_df[exec_col], 
        y=cumulative_df['Cumulative_PL'], 
        mode='lines', 
        name='累計總盈虧', 
        line=dict(color='#2E86AB', width=2)
    ))
    fig.add_trace(go.Scatter(
        x=cumulative_df[exec_col], 
        y=cumulative_df['Scalper_Cumulative_PL'], 
        mode='lines', 
        name=f'Scalper (<{scalper_minutes}分鐘)', 
        line=dict(color='#F39C12', width=2, dash='dot')
    ))
    fig.add_hline(y=0, line_dash="dash", line_color="gray")
    fig.update_layout(
        title='📈 個人累計盈虧', 
        height=350, 
        legend=dict(orientation="h", y=1.05), 
        plot_bgcolor='rgba(248,249,250,1)'
    )
    return fig


def get_client_details(df, aid, initial_balance, scalper_threshold_seconds):
    """獲取客戶詳細資料"""
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    closing_df = filter_closing_trades(df)
    client_df = closing_df[closing_df[aid_col] == str(aid)].copy()
    if client_df.empty:
        return None
    
    net_pl = client_df['Net_PL'].sum()
    trade_count = len(client_df)
    wins = (client_df['Net_PL'] > 0).sum()
    win_rate = (wins / trade_count * 100) if trade_count > 0 else 0
    avg_hold_seconds = client_df['Hold_Seconds'].mean()
    avg_hold_seconds = avg_hold_seconds if pd.notna(avg_hold_seconds) else 0
    
    gains = client_df[client_df[closed_pl_col] > 0][closed_pl_col].sum()
    losses = abs(client_df[client_df[closed_pl_col] < 0][closed_pl_col].sum())
    profit_factor = gains / losses if losses > 0 else (5.0 if gains > 0 else 0)
    
    # Sharpe
    sharpe = client_df['Net_PL'].mean() / client_df['Net_PL'].std() if client_df['Net_PL'].std() > 0 and trade_count >= 3 else 0
    
    aid_sorted = client_df.sort_values(exec_col)
    if len(aid_sorted) >= 2:
        cumulative_pl = aid_sorted['Net_PL'].cumsum()
        equity = initial_balance + cumulative_pl
        running_max = equity.cummax()
        drawdown = np.where(running_max != 0, (equity - running_max) / running_max * 100, 0)
        mdd_pct = abs(np.min(drawdown))
    else:
        mdd_pct = 0.0
    
    client_sorted = client_df.sort_values(exec_col).copy()
    client_sorted['Cumulative_PL'] = client_sorted['Net_PL'].cumsum()
    scalper_mask = client_sorted['Hold_Seconds'] < scalper_threshold_seconds
    client_sorted['Scalper_PL'] = np.where(scalper_mask, client_sorted['Net_PL'], 0)
    client_sorted['Scalper_Cumulative_PL'] = client_sorted['Scalper_PL'].cumsum()
    
    if instrument_col in client_df.columns:
        symbol_dist = client_df.groupby(instrument_col).size().reset_index(name='Count')
        symbol_dist.columns = ['Symbol', 'Count']
    else:
        symbol_dist = pd.DataFrame()
    
    behavioral_stats = calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds)
    
    return {
        'net_pl': net_pl, 
        'trade_count': trade_count, 
        'win_rate': win_rate,
        'avg_hold_seconds': avg_hold_seconds, 
        'profit_factor': profit_factor,
        'sharpe': sharpe, 
        'mdd_pct': mdd_pct,
        'cumulative_df': client_sorted[[exec_col, 'Cumulative_PL', 'Scalper_Cumulative_PL']],
        'symbol_dist': symbol_dist, 
        'client_df': client_df, 
        'behavioral': behavioral_stats
    }


def get_client_ranking(aid_stats_df, aid, metric='Net_PL'):
    """獲取客戶排名"""
    sorted_df = aid_stats_df.sort_values(metric, ascending=False).reset_index(drop=True)
    try:
        rank = sorted_df[sorted_df['AID'] == str(aid)].index[0] + 1
        return rank, len(sorted_df)
    except:
        return None, len(sorted_df)


def export_to_excel(df, aid_stats_df, initial_balance, scalper_threshold_seconds):
    """匯出至 Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment
    output = BytesIO()
    closing_df = filter_closing_trades(df)
    aid_col = COLUMN_MAP['aid']
    
    summary_df = pd.DataFrame([
        ['總交易筆數', len(df)], 
        ['平倉交易筆數', len(closing_df)], 
        ['總客戶數', df[aid_col].nunique()], 
        ['總淨盈虧', round(closing_df['Net_PL'].sum(), 2)], 
        ['初始資金', initial_balance]
    ], columns=['指標', '數值'])
    
    risk_return_df = aid_stats_df[[
        'AID', 'Net_PL', 'MDD_Pct', 'Sharpe', 'Trade_Count', 
        'Win_Rate', 'Profit_Factor', 'Scalper_Ratio', 'Q1', 'Median', 'Q3'
    ]].sort_values('Net_PL', ascending=False)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        risk_return_df.to_excel(writer, sheet_name='Risk_Return', index=False)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
    
    output.seek(0)
    return output


# ==================== 主程式 ====================
def main():
    st.title("📊 交易數據分析系統 v2.4")
    st.markdown("**全局過濾器 | AID 快速交互 | 放大圖表佈局 | Bug 修復版**")
    
    with st.sidebar:
        st.header("⚙️ 全域參數")
        initial_balance = st.number_input(
            "💰 初始資金", 
            value=10000.0,  # 確保為 float
            min_value=0.0,  # 確保為 float
            step=1000.0  # 確保為 float
        )
        scalper_minutes = st.number_input(
            "⏱️ Scalper (分鐘)", 
            value=5.0,  # 確保為 float
            min_value=1.0,  # 確保為 float
            max_value=60.0,  # 確保為 float
            step=1.0  # 確保為 float
        )
        scalper_threshold_seconds = scalper_minutes * 60
        
        st.markdown("---")
        st.header("📁 數據上傳")
        uploaded_files = st.file_uploader(
            "上傳交易數據", 
            type=['xlsx', 'csv'], 
            accept_multiple_files=True
        )
        
        if uploaded_files:
            st.success(f"✅ 已上傳 {len(uploaded_files)} 個檔案")
    
    if not uploaded_files:
        st.info("👈 請在左側上傳交易數據檔案")
        return
    
    with st.spinner("載入數據中..."):
        df = load_and_preprocess(uploaded_files)
    
    if df is None or df.empty:
        st.error("無法載入數據")
        return
    
    display_df = df.copy()
    
    with st.spinner("計算統計中..."):
        aid_stats_df = calculate_all_aid_stats_realtime(
            display_df, 
            initial_balance, 
            scalper_threshold_seconds
        )
    
    st.markdown("---")
    closing_df = filter_closing_trades(display_df)
    aid_col = COLUMN_MAP['aid']
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("總交易筆數", f"{len(display_df):,}")
    col2.metric("平倉交易", f"{len(closing_df):,}")
    col3.metric("客戶數", f"{display_df[aid_col].nunique():,}")
    col4.metric("總淨盈虧", f"${closing_df['Net_PL'].sum():,.2f}")
    
    with st.sidebar:
        st.markdown("---")
        excel_data = export_to_excel(
            display_df, 
            aid_stats_df, 
            initial_balance, 
            scalper_threshold_seconds
        )
        st.download_button(
            "📊 下載 Excel", 
            data=excel_data, 
            file_name=f"report_{datetime.now().strftime('%Y%m%d')}.xlsx", 
            type="primary"
        )
    
    tab1, tab2, tab3 = st.tabs(["📊 整體數據概覽", "👤 個人報告卡", "📅 當日數據概覽"])
    
    # ==================== Tab 1 ====================
    with tab1:
        st.header("📊 整體數據概覽")
        
        # 累計走勢圖
        cumulative_fig, pnl_stats = create_cumulative_pnl_chart(
            display_df, 
            initial_balance, 
            scalper_threshold_seconds
        )
        st.plotly_chart(cumulative_fig, use_container_width=True)
        
        m1, m2 = st.columns(2)
        m1.metric("整體淨盈虧", f"${pnl_stats['total_pnl']:,.2f}")
        m2.metric("Scalper 淨盈虧", f"${pnl_stats['scalper_pnl']:,.2f}")
        
        st.markdown("---")
        
        # 📊 獲利因子 + 交易風格 並排顯示 (st.columns(2))
        st.markdown("### 📊 獲利因子 & 交易風格")
        pf_col, style_col = st.columns(2)
        with pf_col:
            pf_fig, profitable_ratio = create_profit_factor_chart_colored(aid_stats_df)
            st.plotly_chart(pf_fig, use_container_width=True)
            st.success(f"PF > 1.0 佔比: {profitable_ratio:.1f}%")
        with style_col:
            style_pie = create_trading_style_pie(display_df, "🎨 全公司交易風格")
            if style_pie:
                st.plotly_chart(style_pie, use_container_width=True)
        
        st.markdown("---")
        
        # 🎻 小提琴圖 + 統計摘要（放大全寬）
        st.markdown("### 🎻 客戶盈虧分佈")
        violin_fig, violin_stats = create_violin_plot_with_stats(display_df)
        
        # 統計摘要浮動說明欄（左側）+ 圖表（右側）
        stat_col, chart_col = st.columns([1, 3])
        with stat_col:
            st.info(f"""
**📊 統計摘要**
━━━━━━━━━━━━
**客戶總數:** {violin_stats['count']:,}
**盈利客戶:** {violin_stats['profitable']:,} ({violin_stats['profitable']/violin_stats['count']*100:.1f}%)
**虧損客戶:** {violin_stats['losing']:,}
━━━━━━━━━━━━
**平均值:** ${violin_stats['mean']:,.2f}
**中位數:** ${violin_stats['median']:,.2f}
**標準差:** ${violin_stats['std']:,.2f}
━━━━━━━━━━━━
**Q1 (25%):** ${violin_stats['q1']:,.2f}
**Q3 (75%):** ${violin_stats['q3']:,.2f}
**IQR:** ${violin_stats['iqr']:,.2f}
━━━━━━━━━━━━
**異常點:** {violin_stats['outliers']} 位
            """)
            st.markdown("""
**📖 圖例說明**
- 🔵 藍點 = 各 AID
- ⬛ 黑線 = 中位數
- 🔴 紅線 = 平均值
- 📦 白框 = IQR 區間
            """)
        with chart_col:
            st.plotly_chart(violin_fig, use_container_width=True)
        
        st.markdown("---")
        
        # 🎯 風險回報矩陣 (放大全寬)
        st.markdown("### 🎯 風險回報矩陣")
        st.plotly_chart(
            create_risk_return_scatter(aid_stats_df, initial_balance), 
            use_container_width=True
        )
        
        st.markdown("---")
        
        # 每日盈虧
        st.plotly_chart(create_daily_pnl_chart(display_df), use_container_width=True)
        
        st.markdown("---")
        
        # ========== Top 20 盈利英雄榜 ==========
        st.markdown("### 🏆 Top 20 歷史盈利英雄榜")
        st.caption("💡 **點擊表格中的 AID 可選取複製，貼到 Tab 2 搜尋框即可查看詳情**")
        
        min_pnl_h1, min_wr_h1, min_sharpe_h1, max_mdd_h1 = render_global_filters(
            "hist_hero", 0.0, 0.0, -10.0, 100.0
        )
        
        history_hero = calculate_hero_metrics(
            display_df, 
            initial_balance, 
            scalper_threshold_seconds, 
            filter_positive=True, 
            min_pnl=min_pnl_h1, 
            min_winrate=min_wr_h1, 
            min_sharpe=min_sharpe_h1, 
            max_mdd=max_mdd_h1
        )
        
        if not history_hero.empty:
            st.dataframe(
                format_hero_table_display(history_hero), 
                use_container_width=True, 
                hide_index=True, 
                column_config=get_table_column_config()
            )
        else:
            st.info("無符合條件的客戶")
        
        st.markdown("---")
        
        # ========== Top 20 Scalper 英雄榜 ==========
        st.markdown("### 🔥 Top 20 歷史 Scalper 英雄榜")
        
        min_scalp_pct_h, min_scalp_pl_h = render_scalper_filters("hist_scalp", 80.0, 0.0)
        min_pnl_s1, min_wr_s1, min_sharpe_s1, max_mdd_s1 = render_global_filters(
            "hist_scalp_g", 0.0, 0.0, -10.0, 100.0
        )
        
        history_scalp = calculate_hero_metrics(
            display_df, 
            initial_balance, 
            scalper_threshold_seconds,
            filter_positive=True, 
            min_scalp_pct=min_scalp_pct_h, 
            min_scalp_pl=min_scalp_pl_h,
            min_pnl=min_pnl_s1, 
            min_winrate=min_wr_s1, 
            min_sharpe=min_sharpe_s1, 
            max_mdd=max_mdd_s1
        )
        
        if not history_scalp.empty:
            st.dataframe(
                format_hero_table_display(history_scalp), 
                use_container_width=True, 
                hide_index=True, 
                column_config=get_table_column_config()
            )
        else:
            st.info("無符合條件的 Scalper")
    
    # ==================== Tab 2 ====================
    with tab2:
        st.header("👤 個人報告卡")
        
        # 📋 新增提示 caption
        st.caption("📋 **點擊上方表格 AID 複製，在此處貼上 (Ctrl+V) 即可快速搜尋。**")
        
        # AID 輸入框 - 支持文字輸入與貼上
        all_aids = sorted(aid_stats_df['AID'].unique().tolist())
        
        aid_input = st.text_input(
            "🔍 輸入或貼上 AID", 
            value="", 
            placeholder="輸入 AID 數字...", 
            help="直接輸入數字或從表格複製貼上"
        )
        
        # 自動匹配 AID（使用 .strip() 清理空格）
        selected_aid = None
        if aid_input:
            # 🔧 修復：清理輸入（移除空格、逗號等）
            clean_input = aid_input.strip().replace(',', '').replace(' ', '')
            if clean_input in all_aids:
                selected_aid = clean_input
            else:
                # 嘗試模糊匹配
                matches = [a for a in all_aids if clean_input in a]
                if matches:
                    selected_aid = matches[0]
                    st.info(f"自動匹配到: {selected_aid}")
                else:
                    st.warning(f"找不到 AID: {clean_input}")
        
        if selected_aid:
            client_data = get_client_details(
                display_df, 
                selected_aid, 
                initial_balance, 
                scalper_threshold_seconds
            )
            
            if client_data:
                behavioral = client_data['behavioral']
                rank_overall, total_overall = get_client_ranking(
                    aid_stats_df, 
                    selected_aid, 
                    'Net_PL'
                )
                
                st.markdown("---")
                st.markdown(f"## 🆔 AID: {selected_aid}")
                if rank_overall:
                    st.markdown(f"**🏆 整體排名: 第 {rank_overall} 名 / {total_overall} 人**")
                
                # 核心指標
                st.markdown("### 🎯 核心指標")
                c1, c2, c3, c4, c5, c6 = st.columns(6)
                pl_icon = "🟢" if client_data['net_pl'] >= 0 else "🔴"
                c1.metric(f"{pl_icon} 總盈虧", f"${client_data['net_pl']:,.2f}")
                c2.metric("🎯 勝率", f"{client_data['win_rate']:.1f}%")
                c3.metric("📊 PF", f"{client_data['profit_factor']:.2f}")
                c4.metric("📈 Sharpe", f"{client_data['sharpe']:.2f}")
                mdd_icon = "🔴" if client_data['mdd_pct'] > 20 else ""
                c5.metric(f"{mdd_icon}MDD%", f"{client_data['mdd_pct']:.1f}%")
                c6.metric("📝 筆數", f"{client_data['trade_count']}")
                
                # Box Plot 指標
                st.markdown("### 📦 盈虧分佈統計")
                b1, b2, b3, b4 = st.columns(4)
                b1.metric("Q1 (25%)", f"${behavioral['q1']:,.2f}")
                b2.metric("Median", f"${behavioral['median']:,.2f}")
                b3.metric("Q3 (75%)", f"${behavioral['q3']:,.2f}")
                b4.metric("IQR", f"${behavioral['iqr']:,.2f}")
                
                st.markdown("---")
                st.markdown("### ⚔️ 行為分析")
                ba1, ba2 = st.columns(2)
                
                with ba1:
                    st.markdown("#### 多空拆解")
                    st.dataframe(pd.DataFrame({
                        '方向': ['🟢 BUY', '🔴 SELL'],
                        '佔比': [f"{behavioral['buy_ratio']:.1f}%", f"{behavioral['sell_ratio']:.1f}%"],
                        '盈虧': [f"${behavioral['buy_pl']:,.2f}", f"${behavioral['sell_pl']:,.2f}"],
                        '勝率': [f"{behavioral['buy_winrate']:.1f}%", f"{behavioral['sell_winrate']:.1f}%"]
                    }), use_container_width=True, hide_index=True)
                
                with ba2:
                    st.markdown("#### 剝頭皮診斷")
                    st.dataframe(pd.DataFrame({
                        '指標': ['Scalp%', '盈虧貢獻', 'Scalp勝率'],
                        '數值': [
                            f"{behavioral['scalp_ratio']:.1f}%", 
                            f"{behavioral['scalp_contribution']:.1f}%", 
                            f"{behavioral['scalp_winrate']:.1f}%"
                        ]
                    }), use_container_width=True, hide_index=True)
                
                st.markdown("---")
                st.markdown("### 📈 連續紀錄 & 時間效率")
                s1, s2 = st.columns(2)
                with s1:
                    st.dataframe(pd.DataFrame({
                        '類型': ['🏆 連續獲利', '💔 連續虧損'],
                        '次數': [
                            f"{behavioral['max_win_streak']} 次", 
                            f"{behavioral['max_loss_streak']} 次"
                        ],
                        '金額': [
                            f"${behavioral['max_streak_profit']:,.2f}", 
                            f"${behavioral['max_streak_loss']:,.2f}"
                        ]
                    }), use_container_width=True, hide_index=True)
                with s2:
                    st.dataframe(pd.DataFrame({
                        '指標': ['平均持倉', '天數', '分鐘獲利'],
                        '數值': [
                            behavioral['avg_hold_formatted'], 
                            f"{behavioral['avg_hold_days']:.2f}", 
                            f"${behavioral['profit_per_minute']:.4f}"
                        ]
                    }), use_container_width=True, hide_index=True)
                
                # 標籤
                st.markdown("---")
                tags = []
                if behavioral['scalp_ratio'] > 50:
                    tags.append("🔥 高頻型")
                if behavioral['buy_ratio'] > 70:
                    tags.append("⚖️ 偏多")
                elif behavioral['buy_ratio'] < 30:
                    tags.append("⚖️ 偏空")
                if client_data['win_rate'] > 65:
                    tags.append("🎯 高準度")
                if client_data['profit_factor'] > 2:
                    tags.append("💰 高效益")
                if client_data['sharpe'] > 2:
                    tags.append("⭐ 高Sharpe")
                if client_data['mdd_pct'] < 10:
                    tags.append("🛡️ 低風險")
                st.markdown(
                    "**自動標籤:** " + 
                    (" ".join([f"`{t}`" for t in tags]) if tags else "`📊 一般型`")
                )
                
                st.markdown("---")
                ch1, ch2 = st.columns(2)
                with ch1:
                    st.plotly_chart(
                        create_client_cumulative_chart(
                            client_data['cumulative_df'], 
                            scalper_minutes
                        ), 
                        use_container_width=True
                    )
                with ch2:
                    personal_style = create_trading_style_pie(
                        client_data['client_df'], 
                        f"{selected_aid} 風格"
                    )
                    if personal_style:
                        st.plotly_chart(personal_style, use_container_width=True)
            else:
                st.warning(f"找不到 AID: {selected_aid} 的數據")
        else:
            st.info("請輸入或貼上一個 AID 查看報告卡")
    
    # ==================== Tab 3 ====================
    with tab3:
        st.header("📅 當日數據概覽")
        
        exec_col = COLUMN_MAP['execution_time']
        closing_df = filter_closing_trades(display_df)
        latest_date = closing_df[exec_col].dt.date.max()
        st.info(f"📆 分析日期：**{latest_date}**")
        
        day_df = closing_df[closing_df[exec_col].dt.date == latest_date].copy()
        
        if day_df.empty:
            st.warning("當日無交易數據")
        else:
            day_pl = day_df['Net_PL'].sum()
            day_count = len(day_df)
            day_accounts = day_df[aid_col].nunique()
            day_wins = (day_df['Net_PL'] > 0).sum()
            day_wr = (day_wins / day_count * 100) if day_count > 0 else 0
            
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("當日總盈虧", f"${day_pl:,.2f}", delta="盈利" if day_pl >= 0 else "虧損")
            k2.metric("當日交易筆數", f"{day_count:,}")
            k3.metric("當日活躍帳號", f"{day_accounts:,}")
            k4.metric("當日勝率", f"{day_wr:.1f}%")
            
            st.markdown("---")
            
            # 產品分析
            st.markdown("### 📊 當日產品分析")
            profit_products, loss_products = calculate_product_scalp_breakdown(
                day_df, 
                scalper_threshold_seconds
            )
            p1, p2 = st.columns(2)
            with p1:
                profit_chart = create_stacked_product_chart(profit_products, True)
                if profit_chart:
                    st.plotly_chart(profit_chart, use_container_width=True)
                else:
                    st.info("無盈利產品")
            with p2:
                loss_chart = create_stacked_product_chart(loss_products, False)
                if loss_chart:
                    st.plotly_chart(loss_chart, use_container_width=True)
                else:
                    st.info("無虧損產品")
            
            st.markdown("---")
            
            # ========== 當日盈利英雄榜 ==========
            st.markdown("### 🏆 Top 20 當日盈利英雄榜")
            st.caption("💡 **點擊 AID 複製後貼到 Tab 2 搜尋框**")
            
            min_pnl_d1, min_wr_d1, min_sharpe_d1, max_mdd_d1 = render_global_filters(
                "daily_hero", 0.0, 0.0, -10.0, 100.0
            )
            
            daily_hero = calculate_hero_metrics(
                day_df, 
                initial_balance, 
                scalper_threshold_seconds,
                filter_positive=True, 
                min_pnl=min_pnl_d1, 
                min_winrate=min_wr_d1,
                min_sharpe=min_sharpe_d1, 
                max_mdd=max_mdd_d1
            )
            
            if not daily_hero.empty:
                st.dataframe(
                    format_hero_table_display(daily_hero), 
                    use_container_width=True, 
                    hide_index=True, 
                    column_config=get_table_column_config()
                )
                csv_data = daily_hero.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 下載盈利榜 CSV", 
                    data=csv_data, 
                    file_name=f"daily_hero_{latest_date}.csv", 
                    mime="text/csv"
                )
            else:
                st.info("當日無盈利客戶")
            
            st.markdown("---")
            
            # ========== 當日 Scalper 英雄榜 ==========
            st.markdown("### 🔥 Top 20 當日 Scalper 英雄榜")
            
            min_scalp_pct_d, min_scalp_pl_d = render_scalper_filters("daily_scalp", 80.0, 0.0)
            min_pnl_ds, min_wr_ds, min_sharpe_ds, max_mdd_ds = render_global_filters(
                "daily_scalp_g", 0.0, 0.0, -10.0, 100.0
            )
            
            daily_scalp = calculate_hero_metrics(
                day_df, 
                initial_balance, 
                scalper_threshold_seconds,
                filter_positive=True, 
                min_scalp_pct=min_scalp_pct_d, 
                min_scalp_pl=min_scalp_pl_d,
                min_pnl=min_pnl_ds, 
                min_winrate=min_wr_ds, 
                min_sharpe=min_sharpe_ds, 
                max_mdd=max_mdd_ds
            )
            
            if not daily_scalp.empty:
                st.dataframe(
                    format_hero_table_display(daily_scalp), 
                    use_container_width=True, 
                    hide_index=True, 
                    column_config=get_table_column_config()
                )
                csv_scalp = daily_scalp.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 下載 Scalper 榜 CSV", 
                    data=csv_scalp, 
                    file_name=f"scalper_{latest_date}.csv", 
                    mime="text/csv"
                )
            else:
                st.info("當日無符合條件的 Scalper")


if __name__ == "__main__":
    main()
