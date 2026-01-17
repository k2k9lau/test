import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime

# ==================== 常數定義 (1:1 移植) ====================
COLUMN_MAP = {
    'execution_time': 'Execution Time\n交易时间',
    'open_time': 'Open Time\n开仓时间',
    'aid': 'AID\n用户账号',
    'closed_pl': 'Closed P/L\n平仓盈亏',
    'commission': 'Commission\n手续费',
    'swap': 'Swap\n隔夜利息',
    'instrument': 'Instrument\n交易品种',
    'business_type': 'Business Type\n业务类型',
    'action': 'Action\n交易类型',
    'volume': 'Volume\n开仓数量',
    'side': 'Side\n交易方向'
}


# ==================== 數據載入與預處理 ====================
@st.cache_data(show_spinner=False)
def load_and_preprocess(uploaded_files):
    """載入並預處理交易數據 (1:1 還原 v2.4 邏輯)"""
    dfs = []
    for uploaded_file in uploaded_files:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            dfs.append(df)
        except Exception as e:
            st.error(f"讀取檔案 {uploaded_file.name} 時發生錯誤: {e}")
            continue

    if not dfs:
        return None

    df = pd.concat(dfs, ignore_index=True)
    exec_col = COLUMN_MAP['execution_time']
    if exec_col in df.columns:
        df = df[df[exec_col] != 'Total'].copy()
    df = df.drop_duplicates()

    for col in ['execution_time', 'open_time']:
        if COLUMN_MAP[col] in df.columns:
            df[COLUMN_MAP[col]] = pd.to_datetime(df[COLUMN_MAP[col]], errors='coerce')

    for col in ['closed_pl', 'commission', 'swap']:
        if COLUMN_MAP[col] in df.columns:
            df[COLUMN_MAP[col]] = df[COLUMN_MAP[col]].fillna(0)

    df['Net_PL'] = df[COLUMN_MAP['closed_pl']] + df[COLUMN_MAP['commission']] + df[COLUMN_MAP['swap']]

    exec_time = df[COLUMN_MAP['execution_time']]
    open_time = df[COLUMN_MAP['open_time']]
    df['Hold_Seconds'] = np.where(
        pd.notna(exec_time) & pd.notna(open_time),
        (exec_time - open_time).dt.total_seconds(),
        np.nan
    )
    df['Hold_Minutes'] = df['Hold_Seconds'] / 60

    # 嚴格保留 AID 清洗邏輯
    if COLUMN_MAP['aid'] in df.columns:
        df[COLUMN_MAP['aid']] = (
            df[COLUMN_MAP['aid']]
            .astype(str)
            .str.replace(r'\.0$', '', regex=True)
            .str.replace(',', '', regex=False)
            .str.strip()
        )

    return df


def filter_closing_trades(df):
    """過濾平倉交易"""
    action_col = COLUMN_MAP['action']
    if action_col in df.columns:
        return df[df[action_col] == 'CLOSING'].copy()
    return df


def classify_trading_style(hold_minutes):
    """分類交易風格"""
    if pd.isna(hold_minutes):
        return '短線 (Intraday)'
    elif hold_minutes < 5:
        return '極短線 (Scalp)'
    elif hold_minutes < 60:
        return '短線 (Intraday)'
    elif hold_minutes < 1440:
        return '中線 (Day Trade)'
    else:
        return '長線 (Swing)'


# ==================== 核心計算邏輯 (1:1 移植) ====================

def calculate_hero_metrics(data_df, initial_balance, scalper_threshold_seconds,
                           filter_positive=True, min_scalp_pct=None, min_scalp_pl=None,
                           min_pnl=None, min_winrate=None, min_sharpe=None, max_mdd=None):
    """統一計算英雄榜指標 (包含所有過濾邏輯與 Q1/Q3/IQR/PF 等計算)"""
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']

    closing_df = filter_closing_trades(data_df)
    results = []

    for aid in closing_df[aid_col].unique():
        aid_data = closing_df[closing_df[aid_col] == aid].copy()

        net_pl = aid_data['Net_PL'].sum()
        trade_count = len(aid_data)

        if trade_count == 0:
            continue

        # 篩選條件：僅正盈虧
        if filter_positive and net_pl <= 0:
            continue

        # Scalp 數據
        scalp_trades = aid_data[aid_data['Hold_Seconds'] < scalper_threshold_seconds]
        scalp_count = len(scalp_trades)
        scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
        scalp_pct = (scalp_count / trade_count * 100) if trade_count > 0 else 0

        # 勝率
        wins = (aid_data['Net_PL'] > 0).sum()
        losses = trade_count - wins
        win_rate = (wins / trade_count * 100) if trade_count > 0 else 0

        # Sharpe Ratio
        if trade_count >= 3:
            mean_pl = aid_data['Net_PL'].mean()
            std_pl = aid_data['Net_PL'].std()
            sharpe = mean_pl / std_pl if std_pl > 0 else 0.0
        else:
            sharpe = 0.0

        # MDD% 計算
        aid_sorted = aid_data.sort_values(exec_col)
        if len(aid_sorted) >= 2:
            cumulative_pl = aid_sorted['Net_PL'].cumsum()
            equity = initial_balance + cumulative_pl
            running_max = equity.cummax()
            drawdown = np.where(running_max != 0, (equity - running_max) / running_max * 100, 0)
            mdd_pct = abs(np.min(drawdown))
            max_dd_abs = abs((equity - running_max).min())
        else:
            mdd_pct = 0.0
            max_dd_abs = 0.0

        # 應用過濾器
        if min_scalp_pct is not None and scalp_pct < float(min_scalp_pct):
            continue
        if min_scalp_pl is not None and scalp_pl < float(min_scalp_pl):
            continue
        if min_pnl is not None and net_pl < float(min_pnl):
            continue
        if min_winrate is not None and win_rate < float(min_winrate):
            continue
        if min_sharpe is not None and sharpe < float(min_sharpe):
            continue
        if max_mdd is not None and mdd_pct > float(max_mdd):
            continue

        # Box Plot 指標
        q1 = aid_data['Net_PL'].quantile(0.25)
        median = aid_data['Net_PL'].median()
        q3 = aid_data['Net_PL'].quantile(0.75)
        iqr = q3 - q1

        # Profit Expectancy
        win_trades = aid_data[aid_data['Net_PL'] > 0]['Net_PL']
        loss_trades = aid_data[aid_data['Net_PL'] < 0]['Net_PL']
        avg_win = win_trades.mean() if len(win_trades) > 0 else 0
        avg_loss = abs(loss_trades.mean()) if len(loss_trades) > 0 else 0
        win_prob = wins / trade_count if trade_count > 0 else 0
        loss_prob = losses / trade_count if trade_count > 0 else 0
        p_exp = (win_prob * avg_win) - (loss_prob * avg_loss)

        # Profit Factor
        gains = win_trades.sum() if len(win_trades) > 0 else 0
        total_losses = abs(loss_trades.sum()) if len(loss_trades) > 0 else 0
        pf = gains / total_losses if total_losses > 0 else (5.0 if gains > 0 else 0.0)

        # Recovery Factor
        rec_f = net_pl / max_dd_abs if max_dd_abs > 0 else (net_pl if net_pl > 0 else 0.0)

        results.append({
            'AID': str(aid),
            '盈虧': round(net_pl, 2),
            'Scalp盈虧': round(scalp_pl, 2),
            'Scalp%': round(scalp_pct, 2),
            'Sharpe': round(sharpe, 2),
            'MDD%': round(mdd_pct, 2),
            'Q1': round(q1, 2),
            'Median': round(median, 2),
            'Q3': round(q3, 2),
            'IQR': round(iqr, 2),
            'P. Exp': round(p_exp, 2),
            'PF': round(pf, 2),
            'Rec.F': round(rec_f, 2),
            '勝率%': round(win_rate, 2),
            '筆數': trade_count
        })

    result_df = pd.DataFrame(results)
    if not result_df.empty:
        result_df = result_df.sort_values('盈虧', ascending=False).head(20)
    return result_df


def calculate_product_scalp_breakdown(day_df, scalper_threshold_seconds):
    """計算產品 Scalp 拆解"""
    instrument_col = COLUMN_MAP['instrument']
    closing_df = filter_closing_trades(day_df)

    if instrument_col not in closing_df.columns:
        return None, None

    results = []
    for product in closing_df[instrument_col].unique():
        prod_data = closing_df[closing_df[instrument_col] == product]
        total_pl = prod_data['Net_PL'].sum()
        scalp_trades = prod_data[prod_data['Hold_Seconds'] < scalper_threshold_seconds]
        non_scalp_trades = prod_data[prod_data['Hold_Seconds'] >= scalper_threshold_seconds]
        scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
        non_scalp_pl = non_scalp_trades['Net_PL'].sum() if not non_scalp_trades.empty else 0
        scalp_pct = (len(scalp_trades) / len(prod_data) * 100) if len(prod_data) > 0 else 0

        results.append({
            'Product': product,
            'Total_PL': total_pl,
            'Scalp_PL': scalp_pl,
            'NonScalp_PL': non_scalp_pl,
            'Scalp_Pct': scalp_pct
        })

    result_df = pd.DataFrame(results)
    if result_df.empty:
        return None, None

    profit_products = result_df[result_df['Total_PL'] > 0].nlargest(5, 'Total_PL')
    loss_products = result_df[result_df['Total_PL'] < 0].nsmallest(5, 'Total_PL')

    return profit_products, loss_products


def calculate_all_aid_stats_realtime(df, initial_balance, scalper_threshold_seconds):
    """計算所有 AID 的即時統計 (用於 Risk Return Scatter 等)"""
    aid_col = COLUMN_MAP['aid']
    volume_col = COLUMN_MAP['volume']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']

    closing_df = filter_closing_trades(df)
    results = []

    for aid in closing_df[aid_col].unique():
        aid_data = closing_df[closing_df[aid_col] == aid].copy()

        net_pl = aid_data['Net_PL'].sum()
        trade_count = len(aid_data)
        trade_volume = aid_data[volume_col].sum() if volume_col in aid_data.columns else trade_count

        wins = (aid_data['Net_PL'] > 0).sum()
        win_rate = (wins / trade_count * 100) if trade_count > 0 else 0

        avg_hold_seconds = aid_data['Hold_Seconds'].mean() if 'Hold_Seconds' in aid_data.columns else 0
        avg_hold_seconds = avg_hold_seconds if pd.notna(avg_hold_seconds) else 0

        scalper_trades = aid_data[aid_data['Hold_Seconds'] < scalper_threshold_seconds]
        scalper_count = len(scalper_trades)
        scalper_ratio = (scalper_count / trade_count * 100) if trade_count > 0 else 0
        scalper_pl = scalper_trades['Net_PL'].sum() if not scalper_trades.empty else 0

        q1 = aid_data['Net_PL'].quantile(0.25)
        median = aid_data['Net_PL'].median()
        q3 = aid_data['Net_PL'].quantile(0.75)

        # Sharpe
        if trade_count >= 3:
            sharpe = aid_data['Net_PL'].mean() / aid_data['Net_PL'].std() if aid_data['Net_PL'].std() > 0 else 0
        else:
            sharpe = 0

        aid_sorted = aid_data.sort_values(exec_col)
        if len(aid_sorted) >= 2:
            cumulative_pl = aid_sorted['Net_PL'].cumsum()
            equity = initial_balance + cumulative_pl
            running_max = equity.cummax()
            drawdown = np.where(running_max != 0, (equity - running_max) / running_max, 0)
            mdd_pct = abs(np.min(drawdown) * 100)
        else:
            mdd_pct = 0.0

        gains = aid_data[aid_data[closed_pl_col] > 0][closed_pl_col].sum()
        losses = abs(aid_data[aid_data[closed_pl_col] < 0][closed_pl_col].sum())
        profit_factor = gains / losses if losses > 0 else (5.0 if gains > 0 else 0)

        if instrument_col in aid_data.columns and not aid_data[instrument_col].empty:
            main_symbol = aid_data[instrument_col].mode().iloc[0] if len(aid_data[instrument_col].mode()) > 0 else 'N/A'
        else:
            main_symbol = 'N/A'

        results.append({
            'AID': str(aid),
            'Net_PL': round(net_pl, 2),
            'Trade_Count': trade_count,
            'Trade_Volume': round(trade_volume, 2),
            'Win_Rate': round(win_rate, 2),
            'Avg_Hold_Seconds': round(avg_hold_seconds, 2),
            'MDD_Pct': round(mdd_pct, 2),
            'Profit_Factor': round(profit_factor, 2),
            'Scalper_Count': scalper_count,
            'Scalper_Ratio': round(scalper_ratio, 2),
            'Scalper_PL': round(scalper_pl, 2),
            'Sharpe': round(sharpe, 2),
            'Q1': round(q1, 2),
            'Median': round(median, 2),
            'Q3': round(q3, 2),
            'Main_Symbol': main_symbol
        })

    return pd.DataFrame(results)


def calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds):
    """計算深度行為統計"""
    side_col = COLUMN_MAP['side']

    total_trades = len(client_df)
    total_pl = client_df['Net_PL'].sum()
    total_minutes = client_df['Hold_Minutes'].sum() if 'Hold_Minutes' in client_df.columns else 0
    total_minutes = total_minutes if pd.notna(total_minutes) else 0

    pnl_signs = (client_df['Net_PL'] > 0).astype(int)
    streaks = []
    current_streak = 1
    current_type = pnl_signs.iloc[0] if len(pnl_signs) > 0 else 0

    for i in range(1, len(pnl_signs)):
        if pnl_signs.iloc[i] == current_type:
            current_streak += 1
        else:
            streaks.append((current_type, current_streak))
            current_streak = 1
            current_type = pnl_signs.iloc[i]
    if len(pnl_signs) > 0:
        streaks.append((current_type, current_streak))

    win_streaks = [s[1] for s in streaks if s[0] == 1]
    loss_streaks = [s[1] for s in streaks if s[0] == 0]
    max_win_streak = max(win_streaks) if win_streaks else 0
    max_loss_streak = max(loss_streaks) if loss_streaks else 0

    client_sorted = client_df.sort_values(COLUMN_MAP['execution_time']).copy()
    client_sorted['streak_group'] = (client_sorted['Net_PL'] > 0).ne((client_sorted['Net_PL'] > 0).shift()).cumsum()
    streak_sums = client_sorted.groupby('streak_group')['Net_PL'].sum()
    max_streak_profit = streak_sums.max() if not streak_sums.empty else 0
    max_streak_loss = streak_sums.min() if not streak_sums.empty else 0

    buy_trades = client_df[client_df[side_col] == 'BUY'] if side_col in client_df.columns else pd.DataFrame()
    sell_trades = client_df[client_df[side_col] == 'SELL'] if side_col in client_df.columns else pd.DataFrame()

    buy_count, sell_count = len(buy_trades), len(sell_trades)
    buy_ratio = (buy_count / total_trades * 100) if total_trades > 0 else 0
    sell_ratio = (sell_count / total_trades * 100) if total_trades > 0 else 0
    buy_pl = buy_trades['Net_PL'].sum() if not buy_trades.empty else 0
    sell_pl = sell_trades['Net_PL'].sum() if not sell_trades.empty else 0
    buy_wins = (buy_trades['Net_PL'] > 0).sum() if not buy_trades.empty else 0
    sell_wins = (sell_trades['Net_PL'] > 0).sum() if not sell_trades.empty else 0
    buy_winrate = (buy_wins / buy_count * 100) if buy_count > 0 else 0
    sell_winrate = (sell_wins / sell_count * 100) if sell_count > 0 else 0

    scalp_trades = client_df[client_df['Hold_Seconds'] < scalper_threshold_seconds]
    scalp_count = len(scalp_trades)
    scalp_ratio = (scalp_count / total_trades * 100) if total_trades > 0 else 0
    scalp_pl = scalp_trades['Net_PL'].sum() if not scalp_trades.empty else 0
    scalp_contribution = (scalp_pl / total_pl * 100) if total_pl != 0 else 0
    scalp_wins = (scalp_trades['Net_PL'] > 0).sum() if not scalp_trades.empty else 0
    scalp_winrate = (scalp_wins / scalp_count * 100) if scalp_count > 0 else 0

    q1 = client_df['Net_PL'].quantile(0.25)
    median = client_df['Net_PL'].median()
    q3 = client_df['Net_PL'].quantile(0.75)
    iqr = q3 - q1

    avg_minutes = total_minutes / total_trades if total_trades > 0 else 0
    profit_per_minute = total_pl / total_minutes if total_minutes > 0 else 0
    avg_seconds = avg_minutes * 60
    hours = int(avg_seconds // 3600)
    minutes = int((avg_seconds % 3600) // 60)
    seconds = int(avg_seconds % 60)
    avg_hold_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    avg_hold_days = avg_minutes / 1440

    return {
        'max_win_streak': max_win_streak,
        'max_loss_streak': max_loss_streak,
        'max_streak_profit': max_streak_profit,
        'max_streak_loss': max_streak_loss,
        'buy_count': buy_count,
        'sell_count': sell_count,
        'buy_ratio': buy_ratio,
        'sell_ratio': sell_ratio,
        'buy_pl': buy_pl,
        'sell_pl': sell_pl,
        'buy_winrate': buy_winrate,
        'sell_winrate': sell_winrate,
        'scalp_count': scalp_count,
        'scalp_ratio': scalp_ratio,
        'scalp_pl': scalp_pl,
        'scalp_contribution': scalp_contribution,
        'scalp_winrate': scalp_winrate,
        'avg_hold_formatted': avg_hold_formatted,
        'avg_hold_days': avg_hold_days,
        'profit_per_minute': profit_per_minute,
        'q1': q1,
        'median': median,
        'q3': q3,
        'iqr': iqr
    }


def get_client_details(df, aid, initial_balance, scalper_threshold_seconds):
    """獲取客戶詳細資料 (Tab 2 用)"""
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']

    closing_df = filter_closing_trades(df)
    client_df = closing_df[closing_df[aid_col] == str(aid)].copy()
    if client_df.empty:
        return None

    net_pl = client_df['Net_PL'].sum()
    trade_count = len(client_df)
    wins = (client_df['Net_PL'] > 0).sum()
    win_rate = (wins / trade_count * 100) if trade_count > 0 else 0
    avg_hold_seconds = client_df['Hold_Seconds'].mean()
    avg_hold_seconds = avg_hold_seconds if pd.notna(avg_hold_seconds) else 0

    gains = client_df[client_df[closed_pl_col] > 0][closed_pl_col].sum()
    losses = abs(client_df[client_df[closed_pl_col] < 0][closed_pl_col].sum())
    profit_factor = gains / losses if losses > 0 else (5.0 if gains > 0 else 0)

    # Sharpe
    sharpe = client_df['Net_PL'].mean() / client_df['Net_PL'].std() if client_df[
                                                                           'Net_PL'].std() > 0 and trade_count >= 3 else 0

    aid_sorted = client_df.sort_values(exec_col)
    if len(aid_sorted) >= 2:
        cumulative_pl = aid_sorted['Net_PL'].cumsum()
        equity = initial_balance + cumulative_pl
        running_max = equity.cummax()
        drawdown = np.where(running_max != 0, (equity - running_max) / running_max * 100, 0)
        mdd_pct = abs(np.min(drawdown))
    else:
        mdd_pct = 0.0

    client_sorted = client_df.sort_values(exec_col).copy()
    client_sorted['Cumulative_PL'] = client_sorted['Net_PL'].cumsum()
    scalper_mask = client_sorted['Hold_Seconds'] < scalper_threshold_seconds
    client_sorted['Scalper_PL'] = np.where(scalper_mask, client_sorted['Net_PL'], 0)
    client_sorted['Scalper_Cumulative_PL'] = client_sorted['Scalper_PL'].cumsum()

    if instrument_col in client_df.columns:
        symbol_dist = client_df.groupby(instrument_col).size().reset_index(name='Count')
        symbol_dist.columns = ['Symbol', 'Count']
    else:
        symbol_dist = pd.DataFrame()

    behavioral_stats = calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds)

    return {
        'net_pl': net_pl,
        'trade_count': trade_count,
        'win_rate': win_rate,
        'avg_hold_seconds': avg_hold_seconds,
        'profit_factor': profit_factor,
        'sharpe': sharpe,
        'mdd_pct': mdd_pct,
        'cumulative_df': client_sorted[[exec_col, 'Cumulative_PL', 'Scalper_Cumulative_PL']],
        'symbol_dist': symbol_dist,
        'client_df': client_df,
        'behavioral': behavioral_stats
    }


def get_client_ranking(aid_stats_df, aid, metric='Net_PL'):
    """獲取客戶排名"""
    sorted_df = aid_stats_df.sort_values(metric, ascending=False).reset_index(drop=True)
    try:
        rank = sorted_df[sorted_df['AID'] == str(aid)].index[0] + 1
        return rank, len(sorted_df)
    except:
        return None, len(sorted_df)


def export_to_excel(df, aid_stats_df, initial_balance, scalper_threshold_seconds):
    """匯出至 Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment
    output = BytesIO()
    closing_df = filter_closing_trades(df)
    aid_col = COLUMN_MAP['aid']

    summary_df = pd.DataFrame([
        ['總交易筆數', len(df)],
        ['平倉交易筆數', len(closing_df)],
        ['總客戶數', df[aid_col].nunique()],
        ['總淨盈虧', round(closing_df['Net_PL'].sum(), 2)],
        ['初始資金', initial_balance]
    ], columns=['指標', '數值'])

    risk_return_df = aid_stats_df[[
        'AID', 'Net_PL', 'MDD_Pct', 'Sharpe', 'Trade_Count',
        'Win_Rate', 'Profit_Factor', 'Scalper_Ratio', 'Q1', 'Median', 'Q3'
    ]].sort_values('Net_PL', ascending=False)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        risk_return_df.to_excel(writer, sheet_name='Risk_Return', index=False)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

    output.seek(0)
    return output