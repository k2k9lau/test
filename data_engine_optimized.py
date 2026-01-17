import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime

# ==================== å¸¸æ•¸å®šç¾© ====================
COLUMN_MAP = {
    'execution_time': 'Execution Time\näº¤æ˜“æ—¶é—´',
    'open_time': 'Open Time\nå¼€ä»“æ—¶é—´',
    'aid': 'AID\nç”¨æˆ·è´¦å·',
    'closed_pl': 'Closed P/L\nå¹³ä»“ç›ˆäº',
    'commission': 'Commission\næ‰‹ç»­è´¹',
    'swap': 'Swap\néš”å¤œåˆ©æ¯',
    'instrument': 'Instrument\näº¤æ˜“å“ç§',
    'business_type': 'Business Type\nä¸šåŠ¡ç±»å‹',
    'action': 'Action\näº¤æ˜“ç±»å‹',
    'volume': 'Volume\nå¼€ä»“æ•°é‡',
    'side': 'Side\näº¤æ˜“æ–¹å‘'
}


# ==================== æ•¸æ“šè¼‰å…¥èˆ‡é è™•ç† (å„ªåŒ–ç‰ˆ) ====================
@st.cache_data(show_spinner=False, ttl=3600)
def load_and_preprocess(uploaded_files):
    """
    è¼‰å…¥ä¸¦é è™•ç†äº¤æ˜“æ•¸æ“š - å„ªåŒ–è¨˜æ†¶é«”ä½¿ç”¨
    å¢å¼·éŒ¯èª¤è™•ç†å’Œå¤šæª”æ¡ˆç©©å®šæ€§
    """
    if not uploaded_files:
        return None
    
    dfs = []
    error_files = []
    
    for uploaded_file in uploaded_files:
        try:
            # é‡ç½®æª”æ¡ˆæŒ‡é‡
            uploaded_file.seek(0)
            
            if uploaded_file.name.endswith('.csv'):
                # ä½¿ç”¨æ›´é«˜æ•ˆçš„ CSV è®€å–åƒæ•¸
                df = pd.read_csv(
                    uploaded_file, 
                    parse_dates=False,
                    low_memory=False
                )
            elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(uploaded_file, engine='openpyxl')
            else:
                st.warning(f"âš ï¸ è·³éä¸æ”¯æ´çš„æª”æ¡ˆæ ¼å¼: {uploaded_file.name}")
                continue
            
            # é©—è­‰å¿…è¦æ¬„ä½
            required_cols = [COLUMN_MAP['aid'], COLUMN_MAP['execution_time']]
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                error_files.append(f"{uploaded_file.name} (ç¼ºå°‘æ¬„ä½: {', '.join(missing_cols)})")
                continue
            
            dfs.append(df)
            
        except Exception as e:
            error_files.append(f"{uploaded_file.name} ({str(e)})")
            continue
    
    # é¡¯ç¤ºéŒ¯èª¤è¨Šæ¯
    if error_files:
        st.error(f"âŒ ä»¥ä¸‹æª”æ¡ˆè¼‰å…¥å¤±æ•—:\n" + "\n".join([f"- {f}" for f in error_files]))
    
    if not dfs:
        st.error("âŒ ç„¡æ³•è¼‰å…¥ä»»ä½•æœ‰æ•ˆæª”æ¡ˆ")
        return None
    
    # åˆä½µæ•¸æ“šæ¡†
    try:
        df = pd.concat(dfs, ignore_index=True)
    except Exception as e:
        st.error(f"âŒ åˆä½µæª”æ¡ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return None
    
    # æ•¸æ“šæ¸…ç†
    exec_col = COLUMN_MAP['execution_time']
    aid_col = COLUMN_MAP['aid']
    
    try:
        # ç§»é™¤ Total è¡Œ
        if exec_col in df.columns:
            df = df[df[exec_col] != 'Total'].copy()
        
        # å»é‡
        original_count = len(df)
        df = df.drop_duplicates()
        if len(df) < original_count:
            st.info(f"â„¹ï¸ å·²ç§»é™¤ {original_count - len(df)} ç­†é‡è¤‡æ•¸æ“š")
        
        # æ—¥æœŸè½‰æ› - å‘é‡åŒ–è™•ç†
        for col_key in ['execution_time', 'open_time']:
            col_name = COLUMN_MAP[col_key]
            if col_name in df.columns:
                df[col_name] = pd.to_datetime(df[col_name], errors='coerce')
                
                # æª¢æŸ¥ç„¡æ•ˆæ—¥æœŸ
                invalid_dates = df[col_name].isna().sum()
                if invalid_dates > 0:
                    st.warning(f"âš ï¸ {col_name}: {invalid_dates} ç­†æ—¥æœŸç„¡æ•ˆ")
        
        # æ•¸å€¼æ¬„ä½è™•ç† - ä¸€æ¬¡æ€§å¡«å……
        numeric_cols = ['closed_pl', 'commission', 'swap']
        for col_key in numeric_cols:
            col_name = COLUMN_MAP[col_key]
            if col_name in df.columns:
                df[col_name] = pd.to_numeric(df[col_name], errors='coerce').fillna(0)
        
        # å‘é‡åŒ–è¨ˆç®— Net_PL
        df['Net_PL'] = (
            df[COLUMN_MAP['closed_pl']] + 
            df[COLUMN_MAP['commission']] + 
            df[COLUMN_MAP['swap']]
        )
        
        # å‘é‡åŒ–è¨ˆç®—æŒå€‰æ™‚é–“
        exec_time = df[COLUMN_MAP['execution_time']]
        open_time = df[COLUMN_MAP['open_time']]
        
        df['Hold_Seconds'] = np.where(
            pd.notna(exec_time) & pd.notna(open_time),
            (exec_time - open_time).dt.total_seconds(),
            np.nan
        )
        df['Hold_Minutes'] = df['Hold_Seconds'] / 60
        
        # AID æ¸…æ´— - ç¢ºä¿ä¸æœƒè®Šæˆ NaN
        if aid_col in df.columns:
            df[aid_col] = (
                df[aid_col]
                .astype(str)
                .str.replace(r'\.0$', '', regex=True)
                .str.replace(',', '', regex=False)
                .str.strip()
                .replace('nan', '')  # ç§»é™¤å­—ä¸² 'nan'
            )
            
            # ç§»é™¤ç©ºå€¼ AID
            before_filter = len(df)
            df = df[df[aid_col] != ''].copy()
            if len(df) < before_filter:
                st.warning(f"âš ï¸ å·²ç§»é™¤ {before_filter - len(df)} ç­†ç„¡æ•ˆ AID")
            
            # å„ªåŒ–: è½‰æ›ç‚º category ç¯€çœè¨˜æ†¶é«”
            df[aid_col] = df[aid_col].astype('category')
        
        # è¨˜æ†¶é«”å„ªåŒ–: è½‰æ›å…¶ä»–åˆ†é¡æ¬„ä½
        categorical_cols = ['action', 'instrument', 'side', 'business_type']
        for col_key in categorical_cols:
            col_name = COLUMN_MAP[col_key]
            if col_name in df.columns:
                df[col_name] = df[col_name].astype('category')
        
        # è¨˜æ†¶é«”å„ªåŒ–: é™ä½æ•¸å€¼æ¬„ä½ç²¾åº¦ (float64 -> float32)
        float_cols = ['Net_PL', 'Hold_Seconds', 'Hold_Minutes']
        for col_key in numeric_cols:
            if COLUMN_MAP[col_key] in df.columns:
                float_cols.append(COLUMN_MAP[col_key])
        
        for col in float_cols:
            if col in df.columns:
                df[col] = df[col].astype('float32')
        
        if COLUMN_MAP['volume'] in df.columns:
            df[COLUMN_MAP['volume']] = pd.to_numeric(
                df[COLUMN_MAP['volume']], 
                errors='coerce'
            ).fillna(0).astype('float32')
        
        st.success(f"âœ… æˆåŠŸè¼‰å…¥ {len(dfs)} å€‹æª”æ¡ˆï¼Œå…± {len(df):,} ç­†æ•¸æ“š")
        
        return df
        
    except Exception as e:
        st.error(f"âŒ æ•¸æ“šè™•ç†æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return None


def filter_closing_trades(df):
    """éæ¿¾å¹³å€‰äº¤æ˜“"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    action_col = COLUMN_MAP['action']
    if action_col in df.columns:
        return df[df[action_col] == 'CLOSING'].copy()
    return df.copy()


def classify_trading_style(hold_minutes):
    """åˆ†é¡äº¤æ˜“é¢¨æ ¼"""
    if pd.isna(hold_minutes):
        return 'çŸ­ç·š (Intraday)'
    elif hold_minutes < 5:
        return 'æ¥µçŸ­ç·š (Scalp)'
    elif hold_minutes < 60:
        return 'çŸ­ç·š (Intraday)'
    elif hold_minutes < 1440:
        return 'ä¸­ç·š (Day Trade)'
    else:
        return 'é•·ç·š (Swing)'


# ==================== å‘é‡åŒ–è¼”åŠ©å‡½æ•¸ ====================
def calculate_mdd_vectorized(group_df, initial_balance, exec_col):
    """å‘é‡åŒ–è¨ˆç®— MDD% - ä½¿ç”¨ cummax å„ªåŒ–"""
    if len(group_df) < 2:
        return 0.0
    
    try:
        # æŒ‰æ™‚é–“æ’åºä¸¦è¨ˆç®—ç´¯ç©ç›ˆè™§
        sorted_df = group_df.sort_values(exec_col)
        cumulative_pl = sorted_df['Net_PL'].cumsum()
        equity = initial_balance + cumulative_pl
        
        # ä½¿ç”¨ cummax è¨ˆç®—é‹è¡Œæœ€å¤§å€¼ (æ¥µé«˜æ•ˆ)
        running_max = equity.cummax()
        
        # å‘é‡åŒ–è¨ˆç®—å›æ’¤ç™¾åˆ†æ¯”
        drawdown = np.where(running_max != 0, (equity - running_max) / running_max * 100, 0)
        mdd_pct = abs(np.min(drawdown))
        
        return float(mdd_pct)
    except:
        return 0.0


def calculate_sharpe_vectorized(pnl_series, min_trades=3):
    """å‘é‡åŒ–è¨ˆç®— Sharpe Ratio"""
    if len(pnl_series) < min_trades:
        return 0.0
    
    try:
        mean_pl = pnl_series.mean()
        std_pl = pnl_series.std()
        
        if pd.isna(std_pl) or std_pl == 0:
            return 0.0
        
        return float(mean_pl / std_pl)
    except:
        return 0.0


# ==================== æ ¸å¿ƒè¨ˆç®—é‚è¼¯ (å…¨å‘é‡åŒ–ç‰ˆæœ¬) ====================
@st.cache_data(show_spinner=False, ttl=1800)
def calculate_all_aid_stats_realtime(df, initial_balance, scalper_threshold_seconds):
    """
    è¨ˆç®—æ‰€æœ‰ AID çš„å³æ™‚çµ±è¨ˆ - å®Œå…¨å‘é‡åŒ–ç‰ˆæœ¬
    ä½¿ç”¨ groupby + agg æ›¿ä»£å¾ªç’°,å¤§å¹…æå‡æ•ˆèƒ½
    **ä¿®å¾©: æ‰€æœ‰ groupby å¾Œéƒ½åŠ ä¸Š reset_index()**
    """
    if df is None or df.empty:
        return pd.DataFrame()
    
    aid_col = COLUMN_MAP['aid']
    volume_col = COLUMN_MAP['volume']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    closing_df = filter_closing_trades(df)
    
    if closing_df.empty:
        return pd.DataFrame()
    
    try:
        # é å…ˆè¨ˆç®— Scalper mask (å‘é‡åŒ–)
        closing_df = closing_df.copy()
        closing_df['is_scalper'] = closing_df['Hold_Seconds'] < scalper_threshold_seconds
        closing_df['is_win'] = closing_df['Net_PL'] > 0
        closing_df['is_gain'] = closing_df[closed_pl_col] > 0
        closing_df['abs_loss'] = np.where(
            closing_df[closed_pl_col] < 0, 
            abs(closing_df[closed_pl_col]), 
            0
        )
        
        # åŸºç¤çµ±è¨ˆ - ä¸€æ¬¡æ€§ groupby èšåˆ
        basic_stats = closing_df.groupby(aid_col, observed=True).agg({
            'Net_PL': ['sum', 'mean', 'std', 'count'],
            volume_col: 'sum' if volume_col in closing_df.columns else 'count',
            'Hold_Seconds': 'mean',
            'is_win': 'sum',
            'is_scalper': [
                'sum', 
                lambda x: (x * closing_df.loc[x.index, 'Net_PL']).sum()
            ],
            closed_pl_col: [
                lambda x: x[x > 0].sum(),  # gains
                lambda x: abs(x[x < 0].sum())  # losses
            ]
        }).reset_index()  # âœ… ä¿®å¾©: åŠ ä¸Š reset_index()
        
        # å±•å¹³å¤šç´šåˆ—å
        basic_stats.columns = [
            'AID', 'Net_PL', 'Mean_PL', 'Std_PL', 'Trade_Count', 
            'Trade_Volume', 'Avg_Hold_Seconds', 'Wins',
            'Scalper_Count', 'Scalper_PL', 'Gains', 'Losses'
        ]
        
        # è¨ˆç®—ç™¾åˆ†ä½æ•¸ (Q1, Median, Q3) - å‘é‡åŒ–
        quantiles = (
            closing_df.groupby(aid_col, observed=True)['Net_PL']
            .quantile([0.25, 0.5, 0.75])
            .unstack()
            .reset_index()  # âœ… ä¿®å¾©: åŠ ä¸Š reset_index()
        )
        quantiles.columns = ['AID', 'Q1', 'Median', 'Q3']
        
        # åˆä½µåŸºç¤çµ±è¨ˆèˆ‡ç™¾åˆ†ä½æ•¸
        stats = basic_stats.merge(quantiles, on='AID', how='left')
        
        # å‘é‡åŒ–è¨ˆç®—è¡ç”ŸæŒ‡æ¨™
        stats['Win_Rate'] = np.where(
            stats['Trade_Count'] > 0, 
            (stats['Wins'] / stats['Trade_Count'] * 100), 
            0
        )
        
        stats['Scalper_Ratio'] = np.where(
            stats['Trade_Count'] > 0,
            (stats['Scalper_Count'] / stats['Trade_Count'] * 100), 
            0
        )
        
        stats['Profit_Factor'] = np.where(
            stats['Losses'] > 0,
            stats['Gains'] / stats['Losses'],
            np.where(stats['Gains'] > 0, 5.0, 0.0)
        )
        
        stats['IQR'] = stats['Q3'] - stats['Q1']
        
        # Sharpe Ratio (æ‰¹é‡è¨ˆç®—)
        sharpe_values = []
        for aid in stats['AID']:
            aid_data = closing_df[closing_df[aid_col] == aid]['Net_PL']
            sharpe = calculate_sharpe_vectorized(aid_data)
            sharpe_values.append(sharpe)
        stats['Sharpe'] = sharpe_values
        
        # MDD% (æ‰¹é‡è¨ˆç®—)
        mdd_values = []
        for aid in stats['AID']:
            aid_group = closing_df[closing_df[aid_col] == aid]
            mdd = calculate_mdd_vectorized(aid_group, initial_balance, exec_col)
            mdd_values.append(mdd)
        stats['MDD_Pct'] = mdd_values
        
        # ç¢ºä¿æ‰€æœ‰æ•¸å€¼æ¬„ä½ç‚º float
        numeric_columns = [
            'Net_PL', 'Mean_PL', 'Std_PL', 'Trade_Volume', 'Avg_Hold_Seconds',
            'Scalper_PL', 'Gains', 'Losses', 'Win_Rate', 'Scalper_Ratio',
            'Profit_Factor', 'Q1', 'Median', 'Q3', 'IQR', 'Sharpe', 'MDD_Pct'
        ]
        
        for col in numeric_columns:
            if col in stats.columns:
                stats[col] = pd.to_numeric(stats[col], errors='coerce').fillna(0)
        
        # ç¢ºä¿æ•´æ•¸æ¬„ä½
        stats['Trade_Count'] = stats['Trade_Count'].astype(int)
        stats['Wins'] = stats['Wins'].astype(int)
        stats['Scalper_Count'] = stats['Scalper_Count'].astype(int)
        
        return stats
        
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—çµ±è¨ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()


@st.cache_data(show_spinner=False, ttl=1800)
def calculate_hero_metrics(
    df, 
    initial_balance, 
    scalper_threshold_seconds,
    filter_positive=True,
    min_pnl=0.0,
    min_winrate=0.0,
    min_sharpe=-10.0,
    max_mdd=100.0,
    min_scalp_pct=0.0,
    min_scalp_pl=0.0
):
    """
    è¨ˆç®—è‹±é›„æ¦œæŒ‡æ¨™
    **ä¿®å¾©: æ‰€æœ‰ groupby å¾Œéƒ½åŠ ä¸Š reset_index()**
    """
    if df is None or df.empty:
        return pd.DataFrame()
    
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    closing_df = filter_closing_trades(df)
    
    if closing_df.empty:
        return pd.DataFrame()
    
    try:
        # é è¨ˆç®— mask
        closing_df = closing_df.copy()
        closing_df['is_scalper'] = closing_df['Hold_Seconds'] < scalper_threshold_seconds
        closing_df['is_win'] = closing_df['Net_PL'] > 0
        
        # èšåˆçµ±è¨ˆ - âœ… åŠ ä¸Š reset_index()
        hero_stats = closing_df.groupby(aid_col, observed=True).agg({
            'Net_PL': ['sum', 'count'],
            'is_win': 'sum',
            'is_scalper': [
                'sum',
                lambda x: (x * closing_df.loc[x.index, 'Net_PL']).sum()
            ],
            closed_pl_col: [
                lambda x: x[x > 0].sum(),
                lambda x: abs(x[x < 0].sum())
            ]
        }).reset_index()  # âœ… ä¿®å¾©
        
        hero_stats.columns = [
            'AID', 'Net_PL', 'Trade_Count', 'Wins', 
            'Scalper_Count', 'Scalper_PL', 'Gains', 'Losses'
        ]
        
        # å‘é‡åŒ–è¨ˆç®—
        hero_stats['Win_Rate'] = np.where(
            hero_stats['Trade_Count'] > 0,
            (hero_stats['Wins'] / hero_stats['Trade_Count'] * 100),
            0
        )
        
        hero_stats['Scalper_Ratio'] = np.where(
            hero_stats['Trade_Count'] > 0,
            (hero_stats['Scalper_Count'] / hero_stats['Trade_Count'] * 100),
            0
        )
        
        hero_stats['Profit_Factor'] = np.where(
            hero_stats['Losses'] > 0,
            hero_stats['Gains'] / hero_stats['Losses'],
            np.where(hero_stats['Gains'] > 0, 5.0, 0.0)
        )
        
        # Sharpe å’Œ MDD
        sharpe_list = []
        mdd_list = []
        
        for aid in hero_stats['AID']:
            aid_data = closing_df[closing_df[aid_col] == aid]
            sharpe = calculate_sharpe_vectorized(aid_data['Net_PL'])
            mdd = calculate_mdd_vectorized(aid_data, initial_balance, exec_col)
            sharpe_list.append(sharpe)
            mdd_list.append(mdd)
        
        hero_stats['Sharpe'] = sharpe_list
        hero_stats['MDD_Pct'] = mdd_list
        
        # ç¯©é¸
        mask = (hero_stats['Net_PL'] > 0) if filter_positive else (hero_stats['Net_PL'] != 0)
        
        mask &= (hero_stats['Net_PL'] >= min_pnl)
        mask &= (hero_stats['Win_Rate'] >= min_winrate)
        mask &= (hero_stats['Sharpe'] >= min_sharpe)
        mask &= (hero_stats['MDD_Pct'] <= max_mdd)
        mask &= (hero_stats['Scalper_Ratio'] >= min_scalp_pct)
        mask &= (hero_stats['Scalper_PL'] >= min_scalp_pl)
        
        result = hero_stats[mask].copy()
        result = result.sort_values('Net_PL', ascending=False).head(20)
        
        # æ•¸å€¼æ ¼å¼åŒ–
        for col in ['Net_PL', 'Scalper_PL', 'Gains', 'Losses', 'Sharpe']:
            if col in result.columns:
                result[col] = pd.to_numeric(result[col], errors='coerce').fillna(0)
        
        for col in ['Win_Rate', 'Scalper_Ratio', 'Profit_Factor', 'MDD_Pct']:
            if col in result.columns:
                result[col] = pd.to_numeric(result[col], errors='coerce').fillna(0)
        
        # âœ… ä¿®å¾©ï¼šé‡å‘½åæ¬„ä½ä»¥ç¬¦åˆé¡¯ç¤ºå‡½æ•¸çš„æœŸæœ›
        # å°‡å…§éƒ¨ä½¿ç”¨çš„æ¬„ä½åç¨±è½‰æ›ç‚ºé¡¯ç¤ºç”¨çš„æ¬„ä½åç¨±
        result = result.rename(columns={
            'Scalper_Ratio': 'Scalp%',
            'Scalper_PL': 'Scalpç›ˆè™§',
            'Net_PL': 'ç›ˆè™§',
            'Win_Rate': 'å‹ç‡%',
            'MDD_Pct': 'MDD%',
            'Profit_Factor': 'PF'
        })
        
        # è¨ˆç®—é¡å¤–çš„é¡¯ç¤ºæ¬„ä½
        # P.Exp (Profit Expectancy) = (Win_Rate * Avg_Win) - (Loss_Rate * Avg_Loss)
        # ç°¡åŒ–ç‰ˆæœ¬ï¼šä½¿ç”¨ Mean_PL
        if 'ç›ˆè™§' in result.columns and 'Trade_Count' in result.columns:
            result['P. Exp'] = result['ç›ˆè™§'] / result['Trade_Count']
        
        # Rec.F (Recovery Factor) = Net_PL / Max_Drawdown
        if 'ç›ˆè™§' in result.columns and 'MDD%' in result.columns:
            result['Rec.F'] = np.where(
                result['MDD%'] > 0,
                abs(result['ç›ˆè™§'] / result['MDD%']),
                0
            )
        
        # è¨ˆç®— Q1, Median, Q3, IQR (éœ€è¦å¾åŸå§‹æ•¸æ“šé‡æ–°è¨ˆç®—)
        quantile_data = []
        for aid in result['AID']:
            aid_data = closing_df[closing_df[aid_col] == aid]['Net_PL']
            quantile_data.append({
                'AID': aid,
                'Q1': float(aid_data.quantile(0.25)),
                'Median': float(aid_data.median()),
                'Q3': float(aid_data.quantile(0.75))
            })
        
        if quantile_data:
            quantile_df = pd.DataFrame(quantile_data)
            quantile_df['IQR'] = quantile_df['Q3'] - quantile_df['Q1']
            result = result.merge(quantile_df, on='AID', how='left')
        
        # é‡æ–°æ’åºæ¬„ä½
        column_order = [
            'AID', 'ç›ˆè™§', 'Scalpç›ˆè™§', 'Scalp%', 'Sharpe', 'MDD%',
            'Q1', 'Median', 'Q3', 'IQR', 'P. Exp', 'PF', 'Rec.F', 'å‹ç‡%', 'Trade_Count'
        ]
        
        # åªä¿ç•™å­˜åœ¨çš„æ¬„ä½
        existing_cols = [col for col in column_order if col in result.columns]
        result = result[existing_cols]
        
        # é‡å‘½å Trade_Count ç‚ºã€Œç­†æ•¸ã€
        if 'Trade_Count' in result.columns:
            result = result.rename(columns={'Trade_Count': 'ç­†æ•¸'})
        
        return result.reset_index(drop=True)
        
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—è‹±é›„æ¦œæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()


@st.cache_data(show_spinner=False, ttl=1800)
def calculate_product_scalp_breakdown(day_df, scalper_threshold_seconds):
    """
    è¨ˆç®—ç”¢å“ Scalp vs Non-Scalp åˆ†è§£ (Tab 3 ç”¨)
    âœ… ä¿®æ­£ï¼šç¢ºä¿æ¬„ä½åç¨±çµ±ä¸€ç‚º Scalp_PL å’Œ NonScalp_PL
    âœ… ä¿®æ­£ï¼šæ‰€æœ‰ groupby å¾Œéƒ½åŠ ä¸Š reset_index()
    """
    if day_df is None or day_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    instrument_col = COLUMN_MAP['instrument']
    
    if instrument_col not in day_df.columns:
        return pd.DataFrame(), pd.DataFrame()
    
    try:
        day_df = day_df.copy()
        day_df['is_scalper'] = day_df['Hold_Seconds'] < scalper_threshold_seconds
        
        # åˆ†çµ„èšåˆ - âœ… åŠ ä¸Š reset_index()
        product_stats = day_df.groupby(instrument_col, observed=True).agg({
            'Net_PL': 'sum',
            'is_scalper': [
                'sum',
                lambda x: (x * day_df.loc[x.index, 'Net_PL']).sum()
            ]
        }).reset_index()  # âœ… ä¿®å¾©
        
        # âœ… çµ±ä¸€å‘½å - åŠ å¼·é˜²ç¦¦æ€§é‚è¼¯
        product_stats.columns = ['Product', 'Total_PL', 'Scalper_Count', 'Scalp_PL']
        product_stats['NonScalp_PL'] = product_stats['Total_PL'] - product_stats['Scalp_PL']
        
        # ğŸ›¡ï¸ å¼·åˆ¶ç¢ºä¿æ¬„ä½å­˜åœ¨ï¼Œä¸è«–æ•¸æ“šå…§å®¹ç‚ºä½•
        for col in ['Scalp_PL', 'NonScalp_PL', 'Total_PL']:
            if col not in product_stats.columns:
                product_stats[col] = 0.0
        
        # ğŸ›¡ï¸ æ¸…ç† Product æ¬„ä½ä¸­å¯èƒ½çš„æ›è¡Œç¬¦è™Ÿ
        if 'Product' in product_stats.columns:
            product_stats['Product'] = product_stats['Product'].astype(str).str.replace('\n', ' ', regex=False).str.strip()
        
        # åˆ†é›¢ç›ˆåˆ©å’Œè™§æ
        profit_products = product_stats[product_stats['Total_PL'] > 0].copy()
        loss_products = product_stats[product_stats['Total_PL'] < 0].copy()
        
        # æ’åº
        profit_products = profit_products.sort_values('Total_PL', ascending=False)
        loss_products = loss_products.sort_values('Total_PL')
        
        # ğŸ” èª¿è©¦ï¼šç¢ºèªæ¬„ä½å­˜åœ¨
        profit_final = profit_products.reset_index(drop=True)
        loss_final = loss_products.reset_index(drop=True)
        
        # ç¢ºä¿åªè¿”å›éœ€è¦çš„æ¬„ä½ï¼Œé¿å…å¤šé¤˜æ¬„ä½å¹²æ“¾
        required_cols = ['Product', 'Scalp_PL', 'NonScalp_PL', 'Total_PL']
        
        if not profit_final.empty:
            profit_final = profit_final[required_cols]
        if not loss_final.empty:
            loss_final = loss_final[required_cols]
        
        return profit_final, loss_final
        
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—ç”¢å“åˆ†è§£æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame(), pd.DataFrame()


def calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds):
    """
    æ·±åº¦è¡Œç‚ºçµ±è¨ˆåˆ†æ
    **ä¿®å¾©: æ‰€æœ‰ groupby å¾Œéƒ½åŠ ä¸Š reset_index()**
    """
    if client_df is None or client_df.empty:
        return {}
    
    side_col = COLUMN_MAP['side']
    total_trades = len(client_df)
    total_pl = float(client_df['Net_PL'].sum())
    total_minutes = float(client_df['Hold_Minutes'].sum())
    
    try:
        # é€£çºŒç´€éŒ„è¨ˆç®— (å‘é‡åŒ–)
        pnl_signs = (client_df['Net_PL'] > 0).astype(int).values
        
        # ä½¿ç”¨ numpy diff æ‰¾å‡ºè®ŠåŒ–é»
        changes = np.concatenate(([0], np.where(np.diff(pnl_signs) != 0)[0] + 1, [len(pnl_signs)]))
        streak_lengths = np.diff(changes)
        streak_types = pnl_signs[changes[:-1]]
        
        win_streaks = streak_lengths[streak_types == 1]
        loss_streaks = streak_lengths[streak_types == 0]
        
        max_win_streak = int(win_streaks.max()) if len(win_streaks) > 0 else 0
        max_loss_streak = int(loss_streaks.max()) if len(loss_streaks) > 0 else 0
        
        # é€£çºŒç›ˆè™§é‡‘é¡
        client_sorted = client_df.sort_values(COLUMN_MAP['execution_time']).copy()
        client_sorted['streak_group'] = (
            client_sorted['Net_PL'] > 0
        ).ne(
            (client_sorted['Net_PL'] > 0).shift()
        ).cumsum()
        
        # âœ… ä¿®å¾©: åŠ ä¸Š reset_index()
        streak_sums = (
            client_sorted.groupby('streak_group', observed=False)['Net_PL']
            .sum()
            .reset_index()
        )
        
        max_streak_profit = float(streak_sums['Net_PL'].max()) if not streak_sums.empty else 0
        max_streak_loss = float(streak_sums['Net_PL'].min()) if not streak_sums.empty else 0
        
        # å¤šç©ºçµ±è¨ˆ (å‘é‡åŒ–)
        if side_col in client_df.columns:
            buy_mask = client_df[side_col] == 'BUY'
            sell_mask = client_df[side_col] == 'SELL'
            
            buy_count = int(buy_mask.sum())
            sell_count = int(sell_mask.sum())
            
            buy_pl = float(client_df.loc[buy_mask, 'Net_PL'].sum()) if buy_count > 0 else 0
            sell_pl = float(client_df.loc[sell_mask, 'Net_PL'].sum()) if sell_count > 0 else 0
            
            buy_wins = int((client_df.loc[buy_mask, 'Net_PL'] > 0).sum()) if buy_count > 0 else 0
            sell_wins = int((client_df.loc[sell_mask, 'Net_PL'] > 0).sum()) if sell_count > 0 else 0
        else:
            buy_count = sell_count = 0
            buy_pl = sell_pl = 0
            buy_wins = sell_wins = 0
        
        buy_ratio = (buy_count / total_trades * 100) if total_trades > 0 else 0
        sell_ratio = (sell_count / total_trades * 100) if total_trades > 0 else 0
        buy_winrate = (buy_wins / buy_count * 100) if buy_count > 0 else 0
        sell_winrate = (sell_wins / sell_count * 100) if sell_count > 0 else 0
        
        # Scalper çµ±è¨ˆ (å‘é‡åŒ–)
        scalp_mask = client_df['Hold_Seconds'] < scalper_threshold_seconds
        scalp_count = int(scalp_mask.sum())
        scalp_pl = float(client_df.loc[scalp_mask, 'Net_PL'].sum()) if scalp_count > 0 else 0
        scalp_wins = int((client_df.loc[scalp_mask, 'Net_PL'] > 0).sum()) if scalp_count > 0 else 0
        
        scalp_ratio = (scalp_count / total_trades * 100) if total_trades > 0 else 0
        scalp_contribution = (scalp_pl / total_pl * 100) if total_pl != 0 else 0
        scalp_winrate = (scalp_wins / scalp_count * 100) if scalp_count > 0 else 0
        
        # Box Plot æŒ‡æ¨™
        q1 = float(client_df['Net_PL'].quantile(0.25))
        median = float(client_df['Net_PL'].median())
        q3 = float(client_df['Net_PL'].quantile(0.75))
        iqr = q3 - q1
        
        # æ™‚é–“æ•ˆç‡
        avg_minutes = total_minutes / total_trades if total_trades > 0 else 0
        profit_per_minute = total_pl / total_minutes if total_minutes > 0 else 0
        avg_seconds = avg_minutes * 60
        
        hours = int(avg_seconds // 3600)
        minutes = int((avg_seconds % 3600) // 60)
        seconds = int(avg_seconds % 60)
        avg_hold_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        avg_hold_days = avg_minutes / 1440
        
        return {
            'max_win_streak': max_win_streak,
            'max_loss_streak': max_loss_streak,
            'max_streak_profit': max_streak_profit,
            'max_streak_loss': max_streak_loss,
            'buy_count': buy_count,
            'sell_count': sell_count,
            'buy_ratio': buy_ratio,
            'sell_ratio': sell_ratio,
            'buy_pl': buy_pl,
            'sell_pl': sell_pl,
            'buy_winrate': buy_winrate,
            'sell_winrate': sell_winrate,
            'scalp_count': scalp_count,
            'scalp_ratio': scalp_ratio,
            'scalp_pl': scalp_pl,
            'scalp_contribution': scalp_contribution,
            'scalp_winrate': scalp_winrate,
            'avg_hold_formatted': avg_hold_formatted,
            'avg_hold_days': avg_hold_days,
            'profit_per_minute': profit_per_minute,
            'q1': q1,
            'median': median,
            'q3': q3,
            'iqr': iqr
        }
        
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—è¡Œç‚ºçµ±è¨ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return {
            'max_win_streak': 0, 'max_loss_streak': 0,
            'max_streak_profit': 0, 'max_streak_loss': 0,
            'buy_count': 0, 'sell_count': 0,
            'buy_ratio': 0, 'sell_ratio': 0,
            'buy_pl': 0, 'sell_pl': 0,
            'buy_winrate': 0, 'sell_winrate': 0,
            'scalp_count': 0, 'scalp_ratio': 0,
            'scalp_pl': 0, 'scalp_contribution': 0,
            'scalp_winrate': 0, 'avg_hold_formatted': '00:00:00',
            'avg_hold_days': 0, 'profit_per_minute': 0,
            'q1': 0, 'median': 0, 'q3': 0, 'iqr': 0
        }


@st.cache_data(show_spinner=False, ttl=1800)
def get_client_details(_df, aid, initial_balance, scalper_threshold_seconds):
    """
    ç²å–å®¢æˆ¶è©³ç´°è³‡æ–™ (Tab 2 ç”¨)
    **ä¿®å¾©: æ‰€æœ‰ groupby å¾Œéƒ½åŠ ä¸Š reset_index()**
    """
    if _df is None or _df.empty:
        return None
    
    aid_col = COLUMN_MAP['aid']
    exec_col = COLUMN_MAP['execution_time']
    instrument_col = COLUMN_MAP['instrument']
    closed_pl_col = COLUMN_MAP['closed_pl']
    
    try:
        closing_df = filter_closing_trades(_df)
        client_df = closing_df[closing_df[aid_col] == str(aid)].copy()
        
        if client_df.empty:
            return None
        
        # åŸºç¤çµ±è¨ˆ
        net_pl = float(client_df['Net_PL'].sum())
        trade_count = len(client_df)
        wins = int((client_df['Net_PL'] > 0).sum())
        win_rate = (wins / trade_count * 100) if trade_count > 0 else 0
        
        avg_hold_seconds = client_df['Hold_Seconds'].mean()
        avg_hold_seconds = float(avg_hold_seconds) if pd.notna(avg_hold_seconds) else 0
        
        # Profit Factor
        gains = float(client_df[client_df[closed_pl_col] > 0][closed_pl_col].sum())
        losses = float(abs(client_df[client_df[closed_pl_col] < 0][closed_pl_col].sum()))
        profit_factor = gains / losses if losses > 0 else (5.0 if gains > 0 else 0)
        
        # Sharpe
        sharpe = calculate_sharpe_vectorized(client_df['Net_PL'])
        
        # MDD
        mdd_pct = calculate_mdd_vectorized(client_df, initial_balance, exec_col)
        
        # ç´¯ç© PL è¨ˆç®—
        client_sorted = client_df.sort_values(exec_col).copy()
        client_sorted['Cumulative_PL'] = client_sorted['Net_PL'].cumsum()
        
        scalper_mask = client_sorted['Hold_Seconds'] < scalper_threshold_seconds
        client_sorted['Scalper_PL'] = np.where(scalper_mask, client_sorted['Net_PL'], 0)
        client_sorted['Scalper_Cumulative_PL'] = client_sorted['Scalper_PL'].cumsum()
        
        # Symbol åˆ†ä½ˆ - âœ… åŠ ä¸Š reset_index()
        if instrument_col in client_df.columns:
            symbol_dist = (
                client_df.groupby(instrument_col, observed=True)
                .size()
                .reset_index(name='Count')  # âœ… ä¿®å¾©
            )
            symbol_dist.columns = ['Symbol', 'Count']
        else:
            symbol_dist = pd.DataFrame()
        
        # è¡Œç‚ºçµ±è¨ˆ
        behavioral_stats = calculate_deep_behavioral_stats(client_df, scalper_threshold_seconds)
        
        return {
            'net_pl': net_pl,
            'trade_count': trade_count,
            'win_rate': win_rate,
            'avg_hold_seconds': avg_hold_seconds,
            'profit_factor': profit_factor,
            'sharpe': sharpe,
            'mdd_pct': mdd_pct,
            'cumulative_df': client_sorted[[exec_col, 'Cumulative_PL', 'Scalper_Cumulative_PL']],
            'symbol_dist': symbol_dist,
            'client_df': client_df,
            'behavioral': behavioral_stats
        }
        
    except Exception as e:
        st.error(f"âŒ ç²å–å®¢æˆ¶è©³æƒ…æ™‚ç™¼ç”ŸéŒ¯èª¤ (AID: {aid}): {e}")
        import traceback
        st.code(traceback.format_exc())
        return None


def get_client_ranking(aid_stats_df, aid, metric='Net_PL'):
    """ç²å–å®¢æˆ¶æ’å"""
    if aid_stats_df is None or aid_stats_df.empty:
        return None, 0
    
    try:
        sorted_df = aid_stats_df.sort_values(metric, ascending=False).reset_index(drop=True)
        rank_df = sorted_df[sorted_df['AID'] == str(aid)]
        
        if rank_df.empty:
            return None, len(sorted_df)
        
        rank = rank_df.index[0] + 1
        return int(rank), len(sorted_df)
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—æ’åæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return None, 0


@st.cache_data(show_spinner=False, ttl=1800)
def export_to_excel(_df, _aid_stats_df, initial_balance, scalper_threshold_seconds):
    """åŒ¯å‡ºè‡³ Excel - å„ªåŒ–ç‰ˆ"""
    if _df is None or _df.empty or _aid_stats_df is None or _aid_stats_df.empty:
        st.warning("âš ï¸ ç„¡æ•¸æ“šå¯åŒ¯å‡º")
        return BytesIO()
    
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        output = BytesIO()
        closing_df = filter_closing_trades(_df)
        aid_col = COLUMN_MAP['aid']
        
        # Summary
        summary_df = pd.DataFrame([
            ['ç¸½äº¤æ˜“ç­†æ•¸', len(_df)],
            ['å¹³å€‰äº¤æ˜“ç­†æ•¸', len(closing_df)],
            ['ç¸½å®¢æˆ¶æ•¸', _df[aid_col].nunique()],
            ['ç¸½æ·¨ç›ˆè™§', round(closing_df['Net_PL'].sum(), 2)],
            ['åˆå§‹è³‡é‡‘', initial_balance]
        ], columns=['æŒ‡æ¨™', 'æ•¸å€¼'])
        
        # Risk Return (å„ªåŒ–: åªé¸å–éœ€è¦çš„æ¬„ä½)
        risk_cols = [
            'AID', 'Net_PL', 'MDD_Pct', 'Sharpe', 'Trade_Count',
            'Win_Rate', 'Profit_Factor', 'Scalper_Ratio', 'Q1', 'Median', 'Q3'
        ]
        
        available_cols = [col for col in risk_cols if col in _aid_stats_df.columns]
        risk_return_df = _aid_stats_df[available_cols].sort_values('Net_PL', ascending=False)
        
        # å¯«å…¥ Excel (å„ªåŒ–: æ¸›å°‘æ ¼å¼è¨­ç½®æ¬¡æ•¸)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            risk_return_df.to_excel(writer, sheet_name='Risk_Return', index=False)
            
            # çµ±ä¸€æ ¼å¼è¨­ç½®
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
            header_align = Alignment(horizontal='center')
            
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                # åªè¨­ç½®ç¬¬ä¸€è¡Œ
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
        
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"âŒ Excel åŒ¯å‡ºæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return BytesIO()


# ==================== æ–°å¢ï¼šå€‹äººç”¢å“ç›ˆè™§åˆ†æ (Tab 2) ====================

@st.cache_data(show_spinner=False, ttl=1800)
def calculate_client_product_breakdown(_client_df, scalper_threshold_seconds):
    """
    è¨ˆç®—å–®ä¸€å®¢æˆ¶çš„ç”¢å“ç´šåˆ¥ç›ˆè™§åˆ†è§£ (å« Scalp åˆ†é¡)
    
    åƒæ•¸:
        _client_df: å–®ä¸€ AID çš„äº¤æ˜“æ•¸æ“š DataFrame
        scalper_threshold_seconds: Scalper ç§’æ•¸é–€æª»
    
    è¿”å›:
        DataFrame with columns: ['Symbol', 'Scalp_PL', 'NonScalp_PL', 'Total_PL']
    """
    if _client_df is None or _client_df.empty:
        return pd.DataFrame()
    
    try:
        instrument_col = COLUMN_MAP['instrument']
        
        # ç¢ºä¿æœ‰ Hold_Seconds æ¬„ä½
        if 'Hold_Seconds' not in _client_df.columns:
            st.warning("âš ï¸ ç¼ºå°‘ Hold_Seconds æ¬„ä½ï¼Œç„¡æ³•åˆ†é¡ Scalp")
            return pd.DataFrame()
        
        df = _client_df.copy()
        
        # å‘é‡åŒ–åˆ†é¡ Scalp
        df['Is_Scalp'] = df['Hold_Seconds'] < scalper_threshold_seconds
        
        # æŒ‰ç”¢å“å’Œ Scalp åˆ†çµ„èšåˆ
        product_agg = (
            df.groupby([instrument_col, 'Is_Scalp'], observed=True)['Net_PL']
            .sum()
            .reset_index()  # âœ… é˜²æ­¢ KeyError
        )
        
        # Pivot è¡¨æ ¼ï¼šè¡Œ=ç”¢å“ï¼Œåˆ—=Scalp/NonScalp
        product_pivot = product_agg.pivot_table(
            index=instrument_col,
            columns='Is_Scalp',
            values='Net_PL',
            fill_value=0
        ).reset_index()  # âœ… é˜²æ­¢ KeyError
        
        # é‡å‘½åæ¬„ä½ - âœ… åŠ å¼·é˜²ç¦¦æ€§é‚è¼¯
        product_pivot.columns.name = None
        
        # ğŸ›¡ï¸ æª¢æŸ¥ä¸¦é‡å‘½å Boolean æ¬„ä½
        current_cols = list(product_pivot.columns)
        
        if True in current_cols and False in current_cols:
            product_pivot = product_pivot.rename(columns={
                True: 'Scalp_PL',
                False: 'NonScalp_PL'
            })
        elif True in current_cols:
            product_pivot = product_pivot.rename(columns={True: 'Scalp_PL'})
            product_pivot['NonScalp_PL'] = 0.0
        elif False in current_cols:
            product_pivot = product_pivot.rename(columns={False: 'NonScalp_PL'})
            product_pivot['Scalp_PL'] = 0.0
        
        # ğŸ›¡ï¸ å¼·åˆ¶ç¢ºä¿æ¬„ä½å­˜åœ¨ï¼Œä¸è«–æ•¸æ“šå…§å®¹ç‚ºä½•
        for col in ['Scalp_PL', 'NonScalp_PL']:
            if col not in product_pivot.columns:
                product_pivot[col] = 0.0
        
        # é‡å‘½åç”¢å“æ¬„ä½ - ğŸ›¡ï¸ æ¸…ç†å¯èƒ½çš„æ›è¡Œç¬¦è™Ÿ
        if instrument_col in product_pivot.columns:
            product_pivot = product_pivot.rename(columns={instrument_col: 'Symbol'})
        else:
            # å˜—è©¦æ‰¾åˆ°åŒ…å« 'Instrument' çš„æ¬„ä½
            for col in product_pivot.columns:
                if 'Instrument' in str(col) or 'äº¤æ˜“å“ç§' in str(col):
                    product_pivot = product_pivot.rename(columns={col: 'Symbol'})
                    break
        
        # è¨ˆç®—ç¸½ç›ˆè™§
        product_pivot['Total_PL'] = product_pivot['Scalp_PL'] + product_pivot['NonScalp_PL']
        
        # æ’åºä¸¦è¿”å›
        product_pivot = product_pivot.sort_values('Total_PL', ascending=False)
        
        return product_pivot[['Symbol', 'Scalp_PL', 'NonScalp_PL', 'Total_PL']]
        
    except Exception as e:
        st.error(f"âŒ è¨ˆç®—ç”¢å“ç›ˆè™§æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()
